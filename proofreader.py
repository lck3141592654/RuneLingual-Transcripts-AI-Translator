#!/usr/bin/env python3
# proofreader.py - automated proofreading core module v1.0
# Four phases: 1a Script Check, 1b Template Match, 2 LLM Eval, 3 Polish, 4 Retry Protect

import sys, os, json, asyncio, time, re, uuid
from pathlib import Path
from datetime import datetime
from copy import deepcopy
import pandas as pd
sys.path.insert(0, str(Path(__file__).parent))

from glossary import load_glossary, auto_extract_glossary, find_term_spans, build_relevance_context
from tm_matcher import match_and_fill
from llm_translator import (
    _group_into_batches, _sanitize_sheet_name,
    save_backup_part, sync_progress, atomic_write_text,
    get_relevant_glossary,
)
from enforcer import (
    scan_issues, _enforce_async,
)
from api_config import API_TIMEOUT, OVER_RETURN_TOLERANCE
from shared_pool import SharedBatchPool

# Constants
CHECKPOINT_SUBDIR = "_proofread_checkpoint"
BACKUP_SUBDIR = "_proofread_backup"
PROGRESS_FILE = "progress.json"
SESSION_FILE = "session.json"
QUICK_CHECKPOINT_SUBDIR = "_quick_checkpoint"

def _timestamp(msg):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"  [{now}] {msg}")

def elapsed(start, label):
    secs = (datetime.now() - start).total_seconds()
    print(f"  [{label}] 耗時: {secs:.1f} 秒")

def _proofread_checkpoint_dir(output_dir=None, sheet_name=None):
    if output_dir:
        base = Path(output_dir) / CHECKPOINT_SUBDIR
    else:
        base = Path(__file__).parent / CHECKPOINT_SUBDIR
    if sheet_name:
        base = base / _sanitize_sheet_name(sheet_name)
    return base

# Session management (isolated from main _checkpoint/)
def _save_session(output_dir, session_info):
    cd = _proofread_checkpoint_dir(output_dir)
    cd.mkdir(parents=True, exist_ok=True)
    tmp = cd / f"{SESSION_FILE}.tmp"
    final = cd / SESSION_FILE
    tmp.write_text(
        json.dumps(session_info, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    with open(tmp, "ab") as sf:
        os.fsync(sf.fileno())
    tmp.replace(final)

def _load_session(output_dir):
    sf = _proofread_checkpoint_dir(output_dir) / SESSION_FILE
    if sf.exists():
        try:
            return json.loads(sf.read_text(encoding="utf-8"))
        except Exception:
            pass
    return None

# Phase checkpoint markers
def _mark_phase_complete(output_dir, sheet_name, phase):
    cd = _proofread_checkpoint_dir(output_dir, sheet_name)
    cd.mkdir(parents=True, exist_ok=True)
    marker = cd / f"_{phase}_done"
    atomic_write_text(marker, json.dumps({ "phase": phase, "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S") }, ensure_ascii=False))

def _is_phase_complete(output_dir, sheet_name, phase):
    marker = _proofread_checkpoint_dir(output_dir, sheet_name) / f"_{phase}_done"
    return marker.exists()


# === Phase 4b: Template Correction ===
TEMPLATE_CHECKPOINT_FILE = "template_indices.json"

def _save_template_checkpoint(output_dir, sheet_name, indices):
    cd = _proofread_checkpoint_dir(output_dir, sheet_name)
    cd.mkdir(parents=True, exist_ok=True)
    atomic_write_text(cd / TEMPLATE_CHECKPOINT_FILE, json.dumps({"template_indices": indices}, ensure_ascii=False))

def template_correction(df, glossary, output_dir, sheet_name=None):
    df = match_and_fill(df, glossary)
    done = list(df.index[df["_status"] == "已處理"])
    if done:
        _save_template_checkpoint(output_dir, sheet_name, done)
        print(f"    模板校正完成：{len(done)} 條已處理")
    if "_status" in df.columns:
        df = df.drop(columns=["_status"])
    return df

# Phase 2 prompt: LLM Dual-round Fluency Evaluation
SYSTEM_PROMPT_EVAL = (
    "你是一个中文翻译品质评估专家。\n"
    "请遵守以下规则：\n"
    "请评估以下译文的流畅度和翻译腔程度，分为四级：\n\n"
    "- 没问题：流畅自然，没有翻译腔\n"
    "- 轻度：稍微不自然，但可理解\n"
    "- 中等：明显翻译腔，需要修改\n"
    "- 严重：完全不符合中文表达习惯\n\n"
    "输出格式为原始 JSON 阵列（不要使用 ```json 代码块包裹），每条物件包含 \"index\"、\"level\" 和 \"reason\" 三个栏位。\n"
    "例如：[{\"index\": 0, \"level\": \"没问题\", \"reason\": \"流畅自然\"}]\n"
    "## 强制术语表（禁止使用官方和社区通常的翻譯，遇到以下英文词必须使用指定的中文翻译）\n"
    "{RELEVANT_GLOSSARY}"
)

def _cross_reference(round1, round2):
    # 跳過組合僅限 (沒問題,沒問題)、(沒問題,輕度)、(輕度,沒問題)；兩輪皆「輕度」仍送潤色
    skip_pairs = {("没问题", "没问题"), ("轻度", "没问题"), ("没问题", "轻度")}
    return (round1.get("level"), round2.get("level")) not in skip_pairs


def _pad_eval_results(results, batch):
    """過濾非 dict 並對回傳條數不足的缺失項用「严重」補位，避免漏報。"""
    clean = [r for r in results if isinstance(r, dict)]
    present = {r.get('index') for r in clean}
    padded = list(clean)
    for item in batch:
        idx = item.get('_idx', 0)
        if idx not in present:
            padded.append({'index': idx, 'level': '严重', 'reason': '評估缺失，預設為嚴重'})
    return padded


async def _evaluate_batch(client, model, batch, glossary, api_id):
    relevant = get_relevant_glossary(batch, glossary)
    glossary_str = "\n".join([f"  {e} → {c}" for e, c in relevant]) if relevant else "  无"
    system_prompt = SYSTEM_PROMPT_EVAL.replace("{RELEVANT_GLOSSARY}", glossary_str)
    user_content = json.dumps([
        {"index": item["_idx"], "english": item["english"], "translation": item["translation"]}
        for item in batch
    ], ensure_ascii=False)

    text = ""
    for attempt in range(3):
        try:
            resp = await client.chat.completions.create(
                model=model,
                                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_content},
                ],

                temperature=0.1,
                timeout=API_TIMEOUT,
            )
            text = resp.choices[0].message.content.strip()
            text = re.sub(r"^```(?:json)?\n?", "", text, flags=re.IGNORECASE)
            text = re.sub(r"\n```$", "", text)
            results = json.loads(text)

            # 方法一：先试明确的 key 名称
            if isinstance(results, dict):
                known_keys = ["evaluations", "results", "data", "assessments"]
                for key in known_keys:
                    if key in results:
                        candidate = results[key]
                        if isinstance(candidate, list) and len(candidate) > 0 and isinstance(candidate[0], dict):
                            results = candidate
                            break

            # 方法二：自动搜寻 index+level+reason 数组
            if isinstance(results, dict):
                found = None
                for key, value in results.items():
                    if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict) \
                            and "index" in value[0] and "level" in value[0]:
                        found = value
                        break
                if found is None:
                    for key, value in results.items():
                        if isinstance(value, dict):
                            for k2, v2 in value.items():
                                if isinstance(v2, list) and len(v2) > 0 and isinstance(v2[0], dict) \
                                        and "index" in v2[0] and "level" in v2[0]:
                                    found = v2
                                    break
                if found is not None:
                    results = found

            # 单一条目：dict 同时有 index+level → 包装为数组
            if isinstance(results, dict) and "index" in results and "level" in results:
                print(f"    [{api_id}] ⚠️ API 回傳單一物件而非陣列（嘗試 {attempt + 1}/3）")
                print(f"    [{api_id}] ⚠️ 如果多次顯示此訊息，代表此 API 不適合翻譯任務，建議從 .env 移除或更換模型")
                if attempt < 2:
                    print(f"    [{api_id}] 準備重試...")
                    raise ValueError("single object returned, need array")
                else:
                    results = [results]
                    print(f"    [{api_id}] 3 次嘗試均回傳單一物件，強制使用")

            if isinstance(results, list) and len(results) > int(len(batch) * OVER_RETURN_TOLERANCE):
                raise ValueError(f"回傳 {len(results)} 條，超過批次 {len(batch)} 的 {int(OVER_RETURN_TOLERANCE * 100)}%，視為失敗重試")
            if isinstance(results, dict):
                results = [results]
            if isinstance(results, list):
                batch_indices = {item.get("_idx", 0) for item in batch}
                mapped = {}
                for r in results:
                    if isinstance(r, dict) and r.get("index") in batch_indices:
                        mapped[r["index"]] = r
                results = list(mapped.values())

            if isinstance(results, list) and len(results) > 0:
                # 低评估率检查（< 75% 視為失敗重試，在所有路徑後執行）
                success_count = sum(1 for r in results if isinstance(r, dict) and r.get("level") and r.get("reason"))
                rate = min(success_count, len(batch)) / len(batch) if len(batch) > 0 else 0
                if rate < 0.75:
                    if attempt < 2:
                        print(f"    [{api_id}] ⚠️ 完成率 {rate:.0%}（{success_count}/{len(batch)}）低於 75%，重試")
                        raise ValueError(f"completion rate {rate:.0%} < 75%")
                    else:
                        print(f"    [{api_id}] ⚠️ 完成率 {rate:.0%}（{success_count}/{len(batch)}）低於 75%，已使用所有嘗試")

                # null 检查
                null_count = sum(1 for r in results if isinstance(r, dict) and not r.get("level"))
                if null_count > len(results) * 0.5:
                    print(f"    [{api_id}] 警告：{null_count}/{len(results)} 条评估为 null")

                # 缺字段检查
                if results and not all(isinstance(r, dict) and "index" in r and "level" in r for r in results):
                    print(f"    [{api_id}] 警告：JSON 陣列中缺少 index/level 字段")
                    print(f"    回傳內容前 200 字: {text[:200]}")

            results = _pad_eval_results(results, batch)
            return results

        except json.JSONDecodeError:
            try:
                from json_repair import repair_json
                repaired = repair_json(text)
                if repaired:
                    results = json.loads(repaired)
                    print(f"    [{api_id}] ⚠️ JSON 修復成功")
                    if isinstance(results, dict):
                        if "evaluations" in results:
                            results = results["evaluations"]
                        elif "index" in results and "level" in results:
                            raise ValueError("single object after repair, need array")
                    if isinstance(results, list) and len(results) > int(len(batch) * OVER_RETURN_TOLERANCE):
                        raise ValueError(f"回傳 {len(results)} 條，超過批次 {len(batch)} 的 {int(OVER_RETURN_TOLERANCE * 100)}%，視為失敗重試")
                    if isinstance(results, list):
                        batch_indices = {item.get("_idx", 0) for item in batch}
                        mapped = {}
                        for r in results:
                            if isinstance(r, dict) and r.get("index") in batch_indices:
                                mapped[r["index"]] = r
                        results = list(mapped.values())
                        success_count = sum(1 for r in results if isinstance(r, dict) and r.get("level") and r.get("reason"))
                        rate = min(success_count, len(batch)) / len(batch) if len(batch) > 0 else 0
                        if rate < 0.75:
                            if attempt < 2:
                                print(f"    [{api_id}] ⚠️ 完成率 {rate:.0%}（{success_count}/{len(batch)}）低於 75%，重試")
                                raise ValueError(f"completion rate {rate:.0%} < 75%")
                            else:
                                print(f"    [{api_id}] ⚠️ 完成率 {rate:.0%}（{success_count}/{len(batch)}）低於 75%，已使用所有嘗試")
                        null_count = sum(1 for r in results if isinstance(r, dict) and not r.get("level"))
                        if null_count > len(results) * 0.5:
                            print(f"    [{api_id}] 警告：{null_count}/{len(results)} 条评估为 null")
                        if results and not all(isinstance(r, dict) and "index" in r and "level" in r for r in results):
                            print(f"    [{api_id}] 警告：JSON 陣列中缺少 index/level 字段")
                            print(f"    回傳內容前 200 字: {text[:200]}")
                        results = _pad_eval_results(results, batch)
                        return results
            except Exception:
                pass
        except ValueError as e:
            print(f"    [{api_id}] API 錯誤 (嘗試 {attempt + 1}/3): {e}")
            if attempt < 2:
                await asyncio.sleep(3)
        except Exception as e:
            if "429" in str(e).lower() or "rate" in str(e).lower():
                raise
            print(f"    [{api_id}] API 錯誤 (嘗試 {attempt + 1}/3): {e}")
            if attempt < 2:
                await asyncio.sleep(3)

    _save_debug_info(api_id, model, len(batch), 0, 0, text[:3000] if 'text' in locals() and text else "All 3 attempts failed")
    return [{"index": item.get("_idx", 0), "level": None, "reason": "評估失敗"} for item in batch]

def _save_phase2_results(output_dir, sheet_name, round1_results, round2_results):
    cd = _proofread_checkpoint_dir(output_dir, sheet_name)
    cd.mkdir(parents=True, exist_ok=True)
    atomic_write_text(cd / "phase2_round1.json", json.dumps(round1_results, ensure_ascii=False))
    atomic_write_text(cd / "phase2_round2.json", json.dumps(round2_results, ensure_ascii=False))
    _mark_phase_complete(output_dir, sheet_name, "phase2")


def _load_phase2_category(output_dir, sheet_name, df):
    cd = _proofread_checkpoint_dir(output_dir, sheet_name)
    f1 = cd / "phase2_round1.json"
    f2 = cd / "phase2_round2.json"
    if not f1.exists() or not f2.exists():
        return []
    try:
        r1 = json.loads(f1.read_text(encoding="utf-8"))
        r2 = json.loads(f2.read_text(encoding="utf-8"))
    except Exception:
        return []
    result = []
    for idx_str, r1r in r1.items():
        idx = int(idx_str)
        r2r = r2.get(idx_str, {"level": "严重", "reason": "评估缺失，预设为严重"})
        if _cross_reference(r1r, r2r):
            if idx in df.index:
                row = df.loc[idx]
                result.append({
                    "index": idx,
                    "english": str(row.get("english", "")),
                    "translation": str(row.get("translation", "")),
                    "round1": r1r,
                    "round2": r2r,
                })
    return result


async def phase2_llm_evaluate(df, glossary, output_dir, sheet_name=None, pool=None):
    pending = []
    for idx, row in df.iterrows():
        trans = row.get("translation")
        if trans is None or (isinstance(trans, float) and pd.isna(trans)):
            continue
        ts = str(trans).strip()
        if not ts or ts.lower() in ("nan", "nat", "none"):
            continue
        pending.append({
            "_idx": idx,
            "english": str(row.get("english", "")),
            "translation": ts,
            "sub_category": str(row.get("sub_category", "")),
        })
    if not pending:
        return df, []
    pending.sort(key=lambda x: (x.get("sub_category", ""), x.get("_idx", 0)))
    batches = _group_into_batches(pending)
    total_items = len(pending)
    print(f"  P2：{total_items} 條待評估，共 {len(batches)} 批")
    if pool is None:
        print("  錯誤：沒有可用的共享 API 池")
        return df, []
    rr1 = {}
    rr2 = {}
    tag = uuid.uuid4().hex[:8]
    total_batches = len(batches)

    # Mix all R1 and R2 batches together in one pool
    all_tasks = []
    for bn, batch in enumerate(batches, 1):
        all_tasks.append((bn, 1, batch))  # (orig_bn, round, data)
        all_tasks.append((bn, 2, batch))

    async def process_batch(ws, job):
        orig_bn, rnd, batch = job.ctx
        cfg = ws["cfg"]
        _timestamp(f"P2-R{rnd} 批次 {orig_bn}/{total_batches} 開始 ({len(batch)} 條) [{cfg.api_id} {cfg.model}]")
        try:
            results = await _evaluate_batch(ws["client"], cfg.model, batch, glossary, cfg.api_id)
            results = results[:len(batch)]
            success_count = sum(1 for r in results if isinstance(r, dict) and r.get("level") and r.get("reason"))
            rr = rr1 if rnd == 1 else rr2
            for res in results:
                idx = res.get("index")
                if idx is not None:
                    rr[str(idx)] = {"index": idx, "level": res.get("level", "没问题"), "reason": res.get("reason", "")}
            part_data = {}
            for res2 in results:
                idx2 = res2.get("index")
                if idx2 is not None:
                    part_data[str(idx2)] = {"index": idx2, "level": res2.get("level", "没问题"), "reason": res2.get("reason", "")}
            if part_data:
                cd2 = _proofread_checkpoint_dir(output_dir, sheet_name)
                cd2.mkdir(parents=True, exist_ok=True)
                atomic_write_text(cd2 / f"{tag}_{orig_bn:06d}_r{rnd}.json", json.dumps(part_data, ensure_ascii=False))
            _timestamp(f"P2-R{rnd} 批次 {orig_bn}/{total_batches} 完成 ({success_count} 條) [{cfg.api_id} {cfg.model}]")
        except Exception as e:
            es = str(e).lower()
            if "429" in es or "rate" in es:
                cfg.mark_429()
                if cfg.is_permanently_disabled:
                    print(f"  [{cfg.api_id}] 429，永久停用")
                else:
                    retries = getattr(cfg, "_429_retries", 0) + 1
                    cfg._429_retries = retries
                    print(f"  [{cfg.api_id}] 429，冷卻 60 秒 (retry #{retries})")
                pool.retry(job)
            else:
                print(f"  [{cfg.api_id}] 錯誤：{e}")
                if hasattr(e, "response") and hasattr(e.response, "text"):
                    try:
                        print(f"    回應: {e.response.text[:500]}")
                    except Exception:
                        pass

    try:
        await asyncio.gather(*[pool.submit(fbn, task, process_batch, ctx=task) for fbn, task in enumerate(all_tasks, 1)])
    except RuntimeError as e:
        if "all APIs permanently disabled" in str(e):
            print("  ⚠️ 所有 API 已永久停用，Phase 2 評估中斷")
            _save_phase2_results(output_dir, sheet_name, rr1, rr2)
            print(f"  P2 完成（部分）：已儲存 {len(rr1)} 條結果")
            return df, []
        raise
    round1_results = deepcopy(rr1)
    round2_results = deepcopy(rr2)
    _save_phase2_results(output_dir, sheet_name, round1_results, round2_results)
    second_category = []
    for idx_str, r1r in round1_results.items():
        idx = int(idx_str)
        r2r = round2_results.get(idx_str, {"level": "严重", "reason": "评估缺失，预设为严重"})
        if _cross_reference(r1r, r2r):
            if idx in df.index:
                row = df.loc[idx]
                second_category.append({
                    "index": idx,
                    "english": str(row.get("english", "")),
                    "translation": str(row.get("translation", "")),
                    "round1": r1r,
                    "round2": r2r,
                })
    print(f"  P2 完成：{len(second_category)} 條第二類問題")
    return df, second_category


# === Phase 3: LLM Polish ===
SYSTEM_PROMPT_POLISH = (
    "你是一个 OSRS（Old School RuneScape）游戏文本的简体中文翻译专家。\n"
    "请对以下原文重新翻译，使译文更加流畅自然，减少翻译腔。\n"
    "请遵守以下规则：\n"
    "1. 只能输出简体中文，除非是占位符、标签或代码\n"
    "2. 保持游戏术语一致\n"
    "3. 保留 HTML/XML 标签（如 <col...>, <br>, </col> 等）\n"
    "4. 保留占位符（如 []内的文字）\n"
    "5. 流畅自然和避免翻译腔优先，不必完全忠实原文内容，但要記住OSRS是西方奇幻類遊戲，不應出現過於中式的字詞\n"
    "6. 每条翻译独立进行，不要合并或省略\n"
    "7. 输出格式为原始 JSON 阵列（不要使用 ```json 代码块包裹），每条物件包含 "
    "\"index\" 和 \"translation\" 两个栏位\n"
    "8. 如果术语表的人名、地名等名詞用'.'分隔，例如'索菲娅.休斯'，你必須使用'.'分隔，不要自行改為'·'或其他方式。"
    "你必須使用'.'分隔，不要自行改為'·'或其他方式。你必須使用'.'分隔，不要自行改為'·'或其他方式。\n"
    "## 强制术语表（禁止使用官方和社区通常的翻譯，遇到以下英文词必须使用指定的中文翻译）\n"
    "{RELEVANT_GLOSSARY}"
)


async def _polish_batch(client, model, batch, glossary, api_id):
    relevant = get_relevant_glossary(batch, glossary)
    glossary_str = "\n".join([f"  {e} \u2192 {c}" for e, c in relevant]) if relevant else "  \u65e0"
    system_prompt = SYSTEM_PROMPT_POLISH.replace("{RELEVANT_GLOSSARY}", glossary_str)
    user_content = json.dumps([
        {"index": item["_idx"], "english": item.get("english", "")}
        for item in batch
    ], ensure_ascii=False)
    text = ""
    for attempt in range(3):
        try:
            resp = await client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_content},
                ],

                temperature=0.1,
                timeout=API_TIMEOUT,
            )
            text = resp.choices[0].message.content.strip()
            text = re.sub(r"^```(?:json)?\n?", "", text, flags=re.IGNORECASE)
            text = re.sub(r"\n```$", "", text)
            results = json.loads(text)
            if isinstance(results, dict):
                known_keys = ["translations", "translated", "data", "results"]
                for key in known_keys:
                    if key in results:
                        candidate = results[key]
                        if isinstance(candidate, list) and len(candidate) > 0 and isinstance(candidate[0], dict):
                            results = candidate
                            break
            if isinstance(results, dict):
                found = None
                for key, value in results.items():
                    if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict) \
                            and "index" in value[0] and "translation" in value[0]:
                        found = value
                        break
                if found is None:
                    for key, value in results.items():
                        if isinstance(value, dict):
                            for k2, v2 in value.items():
                                if isinstance(v2, list) and len(v2) > 0 and isinstance(v2[0], dict) \
                                        and "index" in v2[0] and "translation" in v2[0]:
                                    found = v2
                                    break
                if found is not None:
                    results = found
            if isinstance(results, dict) and "index" in results and "translation" in results:
                print(f"    [{api_id}] \u26a0\ufe0f API \u56de\u50b3\u55ae\u4e00\u7269\u4ef6\u800c\u975e\u9663\u5217\uff08\u5617\u8a66 {attempt + 1}/3\uff09")
                if attempt < 2:
                    print(f"    [{api_id}] \u6e96\u5099\u91cd\u8a66...")
                    raise ValueError("single object returned, need array")
                else:
                    results = [results]
                    print(f"    [{api_id}] 3 \u6b21\u5617\u8a66\u5747\u56de\u50b3\u55ae\u4e00\u7269\u4ef6\uff0c\u5f37\u5236\u4f7f\u7528")
                    print(f"    [{api_id}] \u26a0\ufe0f \u5982\u679c\u591a\u6b21\u986f\u793a\u6b64\u8a0a\u606f\uff0c\u4ee3\u8868\u6b64 API \u4e0d\u9069\u5408\u7ffb\u8b6f\u4efb\u52d9\uff0c\u5efa\u8b70\u5f9e .env \u79fb\u9664\u6216\u66f4\u63db\u6a21\u578b")
            if isinstance(results, list) and len(results) > int(len(batch) * OVER_RETURN_TOLERANCE):
                raise ValueError(f"回傳 {len(results)} 條，超過批次 {len(batch)} 的 {int(OVER_RETURN_TOLERANCE * 100)}%，視為失敗重試")
            if isinstance(results, dict):
                results = [results]
            if isinstance(results, list):
                batch_indices = {item.get("_idx", 0) for item in batch}
                mapped = {}
                for r in results:
                    if isinstance(r, dict) and r.get("index") in batch_indices:
                        mapped[r["index"]] = r
                results = list(mapped.values())

            if isinstance(results, list) and len(results) > 0:
                success = sum(1 for r in results if isinstance(r, dict) and r.get("translation"))
                rate = min(success, len(batch)) / len(batch) if len(batch) > 0 else 0
                if rate < 0.75:
                    if attempt < 2:
                        print(f"    [{api_id}] \u26a0\ufe0f \u5b8c\u6210\u7387 {rate:.0%}\uff08{success}/{len(batch)}\uff09\u4f4e\u65bc 75%\uff0c\u91cd\u8a66")
                        raise ValueError(f"completion rate {rate:.0%} < 75%")
                    else:
                        print(f"    [{api_id}] \u26a0\ufe0f \u5b8c\u6210\u7387 {rate:.0%}\uff08{success}/{len(batch)}\uff09\u4f4e\u65bc 75%\uff0c\u5df2\u4f7f\u7528\u6240\u6709\u5617\u8a66")
                null_count = sum(1 for r in results if isinstance(r, dict) and r.get("translation") is None)
                if null_count > len(results) * 0.5:
                    print(f"    [{api_id}] \u8b66\u544a\uff1a{null_count}/{len(results)} \u689d\u7ffb\u8b6f\u70ba null")
                if results and not all("index" in r and "translation" in r for r in results):
                    print(f"    [{api_id}] \u8b66\u544a\uff1aJSON \u9663\u5217\u4e2d\u7f3a\u5c11 index/translation \u6b04\u4f4d")
                    print(f"    \u56de\u50b3\u5167\u5bb9\u524d 200 \u5b57: {text[:200]}")

            return results
        except json.JSONDecodeError:
            try:
                from json_repair import repair_json
                repaired = repair_json(text)
                if repaired:
                    results = json.loads(repaired)
                    print(f"    [{api_id}] ⚠️ JSON 修復成功")
                    if isinstance(results, dict):
                        if "translations" in results:
                            results = results["translations"]
                        elif "index" in results and "translation" in results:
                            raise ValueError("single object after repair, need array")
                    if isinstance(results, list) and len(results) > int(len(batch) * OVER_RETURN_TOLERANCE):
                        raise ValueError(f"回傳 {len(results)} 條，超過批次 {len(batch)} 的 {int(OVER_RETURN_TOLERANCE * 100)}%，視為失敗重試")
                    if isinstance(results, list):
                        batch_indices = {item.get("_idx", 0) for item in batch}
                        mapped = {}
                        for r in results:
                            if isinstance(r, dict) and r.get("index") in batch_indices:
                                mapped[r["index"]] = r
                        results = list(mapped.values())
                        success = sum(1 for r in results if isinstance(r, dict) and r.get("translation"))
                        rate = min(success, len(batch)) / len(batch) if len(batch) > 0 else 0
                        if rate < 0.75:
                            if attempt < 2:
                                print(f"    [{api_id}] ⚠️ 完成率 {rate:.0%}（{success}/{len(batch)}）低於 75%，重試")
                                raise ValueError(f"completion rate {rate:.0%} < 75%")
                            else:
                                print(f"    [{api_id}] ⚠️ 完成率 {rate:.0%}（{success}/{len(batch)}）低於 75%，已使用所有嘗試")
                        null_count = sum(1 for r in results if isinstance(r, dict) and r.get("translation") is None)
                        if null_count > len(results) * 0.5:
                            print(f"    [{api_id}] 警告：{null_count}/{len(results)} 条翻译为 null")
                        if results and not all("index" in r and "translation" in r for r in results):
                            print(f"    [{api_id}] 警告：JSON 陣列中缺少 index/translation 字段")
                            print(f"    回傳內容前 200 字: {text[:200]}")
                        return results
            except Exception:
                pass
        except ValueError as e:
            print(f"    [{api_id}] \u932f\u8aa4 (\u5617\u8a66 {attempt + 1}/3): {e}")
            if attempt < 2:
                await asyncio.sleep(3)
        except Exception as e:
            if "429" in str(e).lower() or "rate" in str(e).lower():
                raise
            print(f"    [{api_id}] API \u932f\u8aa4 (\u5617\u8a66 {attempt + 1}/3): {e}")
            if attempt < 2:
                await asyncio.sleep(3)
    return [{"index": item.get("_idx", 0), "translation": None, "_error": "Polish failed"} for item in batch]

def _save_polish_checkpoint(output_dir, sheet_name, backup_data):
    cd = _proofread_checkpoint_dir(output_dir, sheet_name)
    cd.mkdir(parents=True, exist_ok=True)
    atomic_write_text(cd / "polish_results.json", json.dumps(backup_data, ensure_ascii=False))
    _mark_phase_complete(output_dir, sheet_name, "polish")


def _apply_polish_from_checkpoint(df, output_dir, sheet_name):
    fp = _proofread_checkpoint_dir(output_dir, sheet_name) / "polish_results.json"
    if fp.exists():
        try:
            data = json.loads(fp.read_text(encoding="utf-8"))
            for idx_str, trans in data.items():
                idx = int(idx_str)
                if idx in df.index:
                    df.at[idx, "translation"] = trans
            print(f"    從 checkpoint 還原 {len(data)} 條潤色結果")
        except Exception:
            pass
    return df


async def polish_translations(df, second_category, glossary, output_dir, sheet_name=None, pool=None):
    all_indices = set()
    for item in second_category:
        all_indices.add(item["index"])
    if not all_indices:
        return df

    if pool is None:
        print("  错误：没有可用的共享 API 池")
        return df

    MAX_ROUNDS = 3
    current_indices = sorted(all_indices)

    for rnd in range(MAX_ROUNDS):
        pending = []
        for idx in current_indices:
            if idx not in df.index:
                continue
            row = df.loc[idx]
            pending.append({
                "_idx": idx,
                "english": str(row.get("english", "")),
                "sub_category": str(row.get("sub_category", "")),
            })
        if not pending:
            break

        pending.sort(key=lambda x: (x.get("sub_category", ""), x.get("_idx", 0)))
        batches = _group_into_batches(pending)
        print(f"  P3 第 {rnd + 1}/{MAX_ROUNDS} 轮：{len(pending)} 条待润色，共 {len(batches)} 批")

        tag = uuid.uuid4().hex[:8]
        all_backup = {}
        total_batches = len(batches)

        async def process_batch(ws, job):
            cfg = ws["cfg"]
            bn, batch = job.batch_num, job.batch
            _timestamp(f"P3 批次 {bn}/{total_batches} 开始 ({len(batch)} 条) [{cfg.api_id} {cfg.model}]")
            try:
                results = await _polish_batch(ws["client"], cfg.model, batch, glossary, cfg.api_id)
                new_backup = {}
                skipped = 0
                for res in results:
                    if not isinstance(res, dict):
                        skipped += 1
                        continue
                    idx = res.get("index")
                    trans = res.get("translation")
                    if idx is not None and trans is not None:
                        if idx in df.index:
                            df.at[idx, "translation"] = trans
                            new_backup[str(idx)] = str(trans)
                            all_backup[str(idx)] = str(trans)
                        else:
                            skipped += 1
                if skipped:
                    print(f"  [{cfg.api_id}] 警告：{skipped} 條回傳無效（非物件或 index 不在工作表中），已忽略")
                if new_backup:
                    save_backup_part(output_dir, tag, bn, new_backup, sheet_name)
                    sync_progress(output_dir, sheet_name)
                _timestamp(f"P3 批次 {bn}/{total_batches} 完成 ({len(new_backup)} 条) [{cfg.api_id} {cfg.model}]")
            except Exception as e:
                es = str(e).lower()
                if "429" in es or "rate" in es:
                    cfg.mark_429()
                    if cfg.is_permanently_disabled:
                        print(f"  [{cfg.api_id}] 429，永久停用")
                    else:
                        retries = getattr(cfg, "_429_retries", 0) + 1
                        cfg._429_retries = retries
                        print(f"  [{cfg.api_id}] 429，冷卻 60 秒 (retry #{retries})")
                    pool.retry(job)
                else:
                    print(f"  [{cfg.api_id}] 错误：{e}")
                    if hasattr(e, "response") and hasattr(e.response, "text"):
                        try:
                            print(f"    回应: {e.response.text[:500]}")
                        except Exception:
                            pass

        try:
            await asyncio.gather(*[pool.submit(bn, b, process_batch) for bn, b in enumerate(batches, 1)])
        except RuntimeError as e:
            if "all APIs permanently disabled" in str(e):
                print("  ⚠️ 所有 API 已永久停用，Phase 3 润色中断")
                if all_backup:
                    _save_polish_checkpoint(output_dir, sheet_name, all_backup)
                    sync_progress(output_dir, sheet_name)
                print(f"  P3 完成（部分）：已储存 {len(all_backup)} 条结果")
                return df
            raise

        if all_backup:
            _save_polish_checkpoint(output_dir, sheet_name, all_backup)
            sync_progress(output_dir, sheet_name)
        else:
            print(f"  P3 警告：所有批次均失败，未储存 checkpoint")

        _timestamp(f"P3 第 {rnd + 1} 轮：{len(all_backup)} 条已润色")

        # After polish, retry items that still have empty/failed translations
        if rnd < MAX_ROUNDS - 1:
            remaining = set()
            for idx in current_indices:
                if idx not in df.index:
                    continue
                row = df.loc[idx]
                trans = row.get("translation")
                # Only retry if translation is missing/failed (quality checks are handled by Phase 4)
                if trans is None or (isinstance(trans, float) and pd.isna(trans)) or str(trans).strip() == "":
                    remaining.add(idx)

            current_indices = sorted(remaining)

            if not current_indices:
                print(f"  P3 第 {rnd + 1} 輪後：全部完成")
                break
            print(f"  P3 第 {rnd + 1} 輪後：{len(current_indices)} 條剩餘")
        else:
            print(f"  P3 完成")

    return df

# === Checkpoint Cleanup ===
def _delete_proofread_checkpoint(output_dir):
    cd = _proofread_checkpoint_dir(output_dir)
    if cd.exists():
        import shutil
        shutil.rmtree(cd)
        print("  校對 checkpoint 已清除：" + str(cd))


def _save_debug_info(api_id, model, batch_size, returned, success, text):
    import uuid
    from datetime import datetime
    debug_dir = Path(__file__).parent / "workplace" / "_debugmessage"
    debug_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    uid = uuid.uuid4().hex[:4]
    (debug_dir / f"debug_{ts}_{uid}.json").write_text(
        json.dumps({
            "api_id": api_id,
            "model": model,
            "batch_size": batch_size,
            "returned": returned,
            "success": success,
            "rate": round(success / batch_size, 3) if batch_size > 0 else 0,
            "response_preview": text[:3000]
        }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"    [{api_id}] \u26a0\ufe0f {success}/{batch_size} \u4f4e\u7ffb\u8bd1\u7387\uff0c\u5df2\u5132\u5b58 debug \u8a0a\u606f")


# === Phase 4: Retry Protect ===
async def retry_protect(df, glossary, output_dir, sheet_name=None, pool=None):
    import re
    import pandas as pd
    # Compute relevant glossary (same logic as enforcer.enforce())
    all_text = " ".join(str(row.get("english", "")).lower() for _, row in df.iterrows())
    _ctx = build_relevance_context(all_text)
    relevant_glossary = {
        e: c for e, c in glossary.items()
        if find_term_spans(e, all_text, _ctx)
    }
    # Run retry directly in current event loop with the shared pool
    if pool is not None:
        df = await _enforce_async(df, relevant_glossary, output_dir=output_dir, sheet_name=sheet_name, shared_pool=pool)
    # Final scan (same as enforcer.enforce())
    final_pool = scan_issues(df, relevant_glossary)
    review_rows = []
    for idx, row, issues in final_pool:
        trans = row.get("translation")
        trans_str = str(trans) if not (trans is None or (isinstance(trans, float) and pd.isna(trans))) else ""
        for issue_type, suggested_fix, desc in issues:
            review_rows.append({
                "english": str(row.get("english", "")),
                "current_translation": trans_str,
                "category": str(row.get("category", "")),
                "issue": f"[{issue_type}] {desc}",
            })
    review_df = pd.DataFrame(review_rows)
    return df, review_df


# === Backup ===
def backup_target_file(excel_path):
    backup_dir = Path(excel_path).parent / BACKUP_SUBDIR
    backup_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{Path(excel_path).stem}_{ts}{Path(excel_path).suffix}"
    backup_path = backup_dir / backup_name
    import shutil
    shutil.copy2(excel_path, backup_path)
    print(f"  備份已建立：{backup_path}")
    return backup_path



def generate_reports(excel_path, all_translated_dfs, all_original_dfs, all_second_category, all_review_rows, output_dir, overall_start):
    from openpyxl.styles import Color, PatternFill
    out_path = Path(output_dir) / f"{Path(excel_path).stem}_proofread_output.xlsx"
    if all_translated_dfs:
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            for s_name, df_s in all_translated_dfs.items():
                df_s.to_excel(writer, index=False, sheet_name=s_name)
        print(f"\n  校對輸出：{out_path}")
    all_report_rows = []
    for sheet_name in all_original_dfs:
        df_orig = all_original_dfs[sheet_name]
        df_pol = all_translated_dfs.get(sheet_name, df_orig)
        for item in all_second_category.get(sheet_name, []):
            idx = item["index"]
            eng = str(df_orig.at[idx, "english"]) if idx in df_orig.index else ""
            orig = str(item.get("translation", ""))
            new = str(df_pol.at[idx, "translation"]) if idx in df_pol.index else orig
            r1 = item["round1"]; r2 = item["round2"]
            detail = f"R1:{r1.get('level','?')}({r1.get('reason','')})|R2:{r2.get('level','?')}({r2.get('reason','')})"
            all_report_rows.append({"sheet_name": sheet_name, "english": eng, "original_translation": orig, "new_translation": new, "remark": f"Type2-{detail}"})
    # Type1：P4a 重譯保護後仍存在的機械問題（術語/佔位符/未翻譯/空格）
    for rdf in all_review_rows:
        for _, row in rdf.iterrows():
            all_report_rows.append({
                "sheet_name": str(row.get("sheet_name", "")),
                "english": str(row.get("english", "")),
                "original_translation": str(row.get("current_translation", "")),
                "new_translation": str(row.get("current_translation", "")),
                "remark": f"Type1-{row.get('issue', '')}",
            })
    if all_report_rows:
        rdf = pd.DataFrame(all_report_rows)
        rcols = ["sheet_name", "english", "original_translation", "new_translation", "remark"]
        rdf = rdf[rcols]
        rp = Path(output_dir) / "proofread_report.xlsx"
        with pd.ExcelWriter(rp, engine="openpyxl") as writer:
            rdf.to_excel(writer, index=False, sheet_name="Sheet1")
            ws = writer.sheets["Sheet1"]
            # Type1: 背景顏色標記（標註問題類型）
            type1_fills = {
                "glossary": PatternFill(start_color=Color(rgb="D9D9D9"), end_color=Color(rgb="D9D9D9"), fill_type="solid"),
                "placeholder": PatternFill(start_color=Color(rgb="BDD7EE"), end_color=Color(rgb="BDD7EE"), fill_type="solid"),
                "untranslated": PatternFill(start_color=Color(rgb="F4CCCC"), end_color=Color(rgb="F4CCCC"), fill_type="solid"),
                "space": PatternFill(start_color=Color(rgb="E2EFDA"), end_color=Color(rgb="E2EFDA"), fill_type="solid"),
            }
            # Type2: 文字顏色標記（Type2-R1 部分藍字，其餘黑字）
            from openpyxl.cell.rich_text import CellRichText, TextBlock
            from openpyxl.cell.text import InlineFont
            f_blue = InlineFont(sz=11, color=Color(rgb="FF0070C0"))
            f_black = InlineFont(sz=11, color=Color(rgb="FF000000"))
            remark_col = None
            for ci, cell in enumerate(ws[1], start=1):
                if cell.value == "remark":
                    remark_col = ci; break
            if remark_col:
                for row in ws.iter_rows(min_row=2):
                    v = str(row[remark_col-1].value) if row[remark_col-1].value else ""
                    if v.startswith("Type1-"):
                        for t, fill in type1_fills.items():
                            if t in v:
                                for cell in row:
                                    cell.fill = fill
                                break
                    elif v.startswith("Type2-R1"):
                        parts = v.split("|R2:", 1)
                        if len(parts) == 2:
                            rich = CellRichText([TextBlock(f_blue, parts[0])])
                            rich += CellRichText([TextBlock(f_black, "|R2:" + parts[1])])
                        else:
                            rich = CellRichText([TextBlock(f_blue, v)])
                        row[remark_col-1].value = rich
                    # 其他 Type2（如 Type2-R2）保留黑色（預設）
        print(f"  proofread_report.xlsx 已寫入：{rp}（{len(rdf)} 條）")
    if all_review_rows:
        combined = pd.concat(all_review_rows, ignore_index=True)
        rvp = Path(output_dir) / "review_report_proofread.xlsx"
        with pd.ExcelWriter(rvp, engine="openpyxl") as writer:
            combined.to_excel(writer, index=False, sheet_name="Sheet1")
            ws = writer.sheets["Sheet1"]
            fg = PatternFill(start_color=Color(rgb="D9D9D9"), end_color=Color(rgb="D9D9D9"), fill_type="solid")
            fp = PatternFill(start_color=Color(rgb="BDD7EE"), end_color=Color(rgb="BDD7EE"), fill_type="solid")
            fu = PatternFill(start_color=Color(rgb="F4CCCC"), end_color=Color(rgb="F4CCCC"), fill_type="solid")
            ic = None
            fs = PatternFill(start_color=Color(rgb="E2EFDA"), end_color=Color(rgb="E2EFDA"), fill_type="solid")
            for ci, cell in enumerate(ws[1], start=1):
                if cell.value == "issue":
                    ic = ci; break
            if ic:
                for row in ws.iter_rows(min_row=2):
                    v = str(row[ic-1].value) if row[ic-1].value else ""
                    fill = fg if v.startswith("[glossary]") else (fp if v.startswith("[placeholder]") else (fu if v.startswith("[untranslated]") else (fs if v.startswith("[space]") else None)))
                    if fill:
                        for cell in row:
                            cell.fill = fill
        print(f"  review_report_proofread.xlsx 已寫入：{rvp}（{len(combined)} 條）")
    else:
        print(f"  review_report：未發現問題")
    cp_dir = _proofread_checkpoint_dir(output_dir)
    if cp_dir.exists():
        import shutil; shutil.rmtree(cp_dir)
        print(f"  已清除校對 checkpoint")
    ecp = Path(output_dir) / "_checkpoint"
    if ecp.exists():
        import shutil; shutil.rmtree(ecp)
        print(f"  已清除強制檢查 checkpoint")
    elapsed(overall_start, "總計")


def _build_session_data(excel_path, glossary_path, glossary_sheets, current_sheet, completed_sheets, pending_sheets, sheet_configs, accumulated_review_rows=None, mode: str = "proofread"):
    session = {
        "excel_path": str(Path(excel_path).resolve()),
        "glossary_path": str(glossary_path) if glossary_path else None,
        "glossary_sheets": glossary_sheets,
        "current_sheet": current_sheet,
        "completed_sheets": list(completed_sheets),
        "pending_sheets": list(pending_sheets),
        "sheet_configs": sheet_configs,
        "mode": mode,
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
    }
    if accumulated_review_rows:
        # 與 batch_translate 一致：存完整內容（dict 清單），而非只存數量
        rows = []
        for df in accumulated_review_rows:
            for _, r in df.iterrows():
                rows.append(r.to_dict())
        session["accumulated_review_rows"] = rows
    return session


def _remove_session_file(output_dir) -> None:
    """完成後清除 session，避免已完成的工作表再次觸發續傳提示。"""
    for sf in (_proofread_checkpoint_dir(output_dir) / SESSION_FILE, _quick_checkpoint_dir(output_dir) / SESSION_FILE):
        if sf.exists():
            try:
                sf.unlink()
            except OSError:
                pass


async def _proofread_phase2(sheet_name, st, glossary, workplace_str, pool):
    df = st["df"]
    _timestamp(f"P2 開始：{sheet_name}...")
    if _is_phase_complete(workplace_str, sheet_name, "phase2"):
        sc = _load_phase2_category(workplace_str, sheet_name, df)
        print(f"    P2 還原：{len(sc)} 條第二類問題")
    else:
        df, sc = await phase2_llm_evaluate(df, glossary, workplace_str, sheet_name=sheet_name, pool=pool)
    st["df"] = df
    st["sc"] = sc


async def _proofread_phase3(sheet_name, st, glossary, workplace_str, pool):
    df = st["df"]
    sc = st["sc"]
    _timestamp(f"P3 開始：{sheet_name}...")
    if _is_phase_complete(workplace_str, sheet_name, "polish"):
        df = _apply_polish_from_checkpoint(df, workplace_str, sheet_name)
    else:
        df = await polish_translations(df, sc, glossary, workplace_str, sheet_name=sheet_name, pool=pool)
    st["df"] = df


async def _proofread_phase4a(sheet_name, st, glossary, workplace_str, pool):
    df = st["df"]
    _timestamp(f"P4a 開始：{sheet_name}...")
    restored = _restore_enforce_backup(df, workplace_str, sheet_name)
    if restored > 0:
        print(f"    從重譯 checkpoint 還原 {restored} 條修正")
    df, review_df = await retry_protect(df, glossary, workplace_str, sheet_name=sheet_name, pool=pool)
    st["df"] = df
    st["review_df"] = review_df


async def run_proofread(excel_path, glossary_path, glossary_sheets, sheet_names, sheet_configs, session=None):
    overall_start = datetime.now()
    workplace = Path(excel_path).parent
    _timestamp("正在備份目標檔案...")
    backup_target_file(excel_path)
    _timestamp("正在載入術語庫...")
    t0 = datetime.now()
    if str(glossary_path) == "__AUTO__":
        glossary = auto_extract_glossary(excel_path, str(workplace))
    else:
        glossary = load_glossary(glossary_path, glossary_sheets)
    elapsed(t0, "術語庫")
    print(f"  術語庫載入完成：{len(glossary)} 條")
    all_translated_dfs = {}
    all_original_dfs = {}
    all_second_category = {}
    all_review_rows = []
    if session and session.get("accumulated_review_rows"):
        all_review_rows.append(pd.DataFrame(session["accumulated_review_rows"]))
        print(f"    還原 {len(session['accumulated_review_rows'])} 條審查記錄")
    completed_sheets = []
    pending_sheets = list(sheet_names)
    _save_session(str(workplace), _build_session_data(
        excel_path, glossary_path, glossary_sheets,
        pending_sheets[0] if pending_sheets else "",
        [], pending_sheets, sheet_configs))
    # Phase barrier: all sheets run each phase in parallel through the shared pool.
    pool = SharedBatchPool()
    await pool.start()
    active = [c for c in pool.api_configs if not c.is_permanently_disabled]
    print(f"  已載入 {len(active)} 個可用 API：")
    for cfg in active:
        print(f"    {cfg.api_id} {cfg.model_provider}/{cfg.model} ({cfg.api_type}, 並發={cfg.parallel_limit})")

    try:
        # Pre-read all worksheets
        sheet_states = {}
        for sheet_name in pending_sheets:
            df_original = pd.read_excel(excel_path, sheet_name=sheet_name, dtype=str)
            all_original_dfs[sheet_name] = df_original
            print(f"    總行數：{len(df_original)}")
            sheet_cfg = sheet_configs.get(sheet_name, {})
            selected_indices = set(sheet_cfg.get("selected_indices", []))
            if _is_phase_complete(str(workplace), sheet_name, "all"):
                if not selected_indices:
                    all_translated_dfs[sheet_name] = df_original
                else:
                    df = df_original[df_original.index.isin(selected_indices)].copy()
                    df = _apply_polish_from_checkpoint(df, str(workplace), sheet_name)
                    restored = _restore_enforce_backup(df, str(workplace), sheet_name)
                    if restored > 0:
                        print(f"    從重譯 checkpoint 還原 {restored} 條修正")
                    df, review_df = await retry_protect(df, glossary, str(workplace), sheet_name=sheet_name, pool=pool)
                    if not review_df.empty:
                        review_df.insert(0, "sheet_name", sheet_name)
                        all_review_rows.append(review_df)
                    df_original.update(df)
                    all_translated_dfs[sheet_name] = df_original
                    sc = _load_phase2_category(str(workplace), sheet_name, df)
                    all_second_category[sheet_name] = sc
                completed_sheets.append(sheet_name)
                print(f"  已從 checkpoint 還原 {sheet_name}")
                continue
            if not selected_indices:
                print(f"  {sheet_name} 未選取任何條目，跳過")
                all_translated_dfs[sheet_name] = df_original
                completed_sheets.append(sheet_name)
                continue
            mask = df_original.index.isin(selected_indices)
            df = df_original[mask].copy()
            sheet_states[sheet_name] = {
                "df_original": df_original, "df": df, "sc": [], "review_df": None,
            }

        # Phase 2: LLM dual-round evaluation (all sheets in parallel)
        t4 = datetime.now()
        _timestamp("Phase 2：LLM 評估（全部工作表並行）...")
        await asyncio.gather(*[
            _proofread_phase2(sheet_name, st, glossary, str(workplace), pool)
            for sheet_name, st in sheet_states.items()
        ])
        elapsed(t4, "P2")

        # Phase 3: LLM polish (all sheets in parallel)
        t5 = datetime.now()
        _timestamp("Phase 3：潤色（全部工作表並行）...")
        await asyncio.gather(*[
            _proofread_phase3(sheet_name, st, glossary, str(workplace), pool)
            for sheet_name, st in sheet_states.items()
        ])
        elapsed(t5, "P3")

        # Phase 4a: retry protect (all sheets in parallel)
        t6 = datetime.now()
        _timestamp("Phase 4a：重譯保護（全部工作表並行）...")
        results = await asyncio.gather(*[
            _proofread_phase4a(sheet_name, st, glossary, str(workplace), pool)
            for sheet_name, st in sheet_states.items()
        ], return_exceptions=True)
        for sheet_name, res in zip(sheet_states, results):
            if isinstance(res, BaseException):
                if isinstance(res, RuntimeError) and "all APIs permanently disabled" in str(res):
                    print("\n  ⚠️所有 API 已永久停用，校對中止")
                raise res
        elapsed(t6, "P4a")

        # Merge results
        for sheet_name, st in sheet_states.items():
            df = st["df"]
            all_second_category[sheet_name] = st["sc"]
            review_df = st["review_df"]
            if review_df is not None and not review_df.empty:
                review_df.insert(0, "sheet_name", sheet_name)
                all_review_rows.append(review_df)
            ecp = Path(str(workplace)) / "_checkpoint" / _sanitize_sheet_name(sheet_name) / "enforce_checkpoint.json"
            if ecp.exists():
                ecp.unlink()
            st["df_original"].update(df)
            all_translated_dfs[sheet_name] = st["df_original"]
            _mark_phase_complete(str(workplace), sheet_name, "all")
            completed_sheets.append(sheet_name)
            print(f"  {sheet_name} 完成")

        pending_sheets.clear()
        _save_session(str(workplace), _build_session_data(
            excel_path, glossary_path, glossary_sheets,
            "", completed_sheets, pending_sheets, sheet_configs,
            accumulated_review_rows=all_review_rows))
    finally:
        await pool.close()

    # Phase 4b：模板校正（在所有重譯完成後執行）
    t_p4b = datetime.now()
    _timestamp("Phase 4b：模板校正...")
    for sheet_name in all_translated_dfs:
        sheet_cfg = sheet_configs.get(sheet_name, {})
        selected_indices = set(sheet_cfg.get("selected_indices", []))
        if not selected_indices:
            continue
        df_original = all_translated_dfs[sheet_name]
        df = df_original[df_original.index.isin(selected_indices)].copy()
        df = template_correction(df, glossary, str(workplace), sheet_name=sheet_name)
        df_original.update(df)
        all_translated_dfs[sheet_name] = df_original
    elapsed(t_p4b, "P4b")

    generate_reports(excel_path, all_translated_dfs, all_original_dfs,
        all_second_category,
        all_review_rows, str(workplace), overall_start)
    _delete_proofread_checkpoint(str(workplace))
    _remove_session_file(str(workplace))


# === 快速校對（重譯模式）===

def _quick_checkpoint_dir(output_dir=None, sheet_name=None):
    if output_dir:
        base = Path(output_dir) / QUICK_CHECKPOINT_SUBDIR
    else:
        base = Path(__file__).parent / QUICK_CHECKPOINT_SUBDIR
    if sheet_name:
        base = base / _sanitize_sheet_name(sheet_name)
    return base


def _save_quick_session(output_dir, session_info):
    cd = _quick_checkpoint_dir(output_dir)
    cd.mkdir(parents=True, exist_ok=True)
    atomic_write_text(cd / SESSION_FILE, json.dumps(session_info, ensure_ascii=False, indent=2))


def _load_quick_session(output_dir):
    sf = _quick_checkpoint_dir(output_dir) / SESSION_FILE
    if sf.exists():
        try:
            return json.loads(sf.read_text(encoding="utf-8"))
        except Exception:
            pass
    return None


def _save_quick_sheet_translations(output_dir, sheet_name, translations):
    cd = _quick_checkpoint_dir(output_dir, sheet_name)
    cd.mkdir(parents=True, exist_ok=True)
    atomic_write_text(cd / "translations.json", json.dumps(translations, ensure_ascii=False))


def _load_quick_sheet_translations(output_dir, sheet_name) -> dict:
    f = _quick_checkpoint_dir(output_dir, sheet_name) / "translations.json"
    if f.exists():
        try:
            return json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _save_quick_sheet_review(output_dir, sheet_name, rows):
    cd = _quick_checkpoint_dir(output_dir, sheet_name)
    cd.mkdir(parents=True, exist_ok=True)
    atomic_write_text(cd / "review_rows.json", json.dumps(rows, ensure_ascii=False))


def _load_quick_sheet_review(output_dir, sheet_name) -> list:
    f = _quick_checkpoint_dir(output_dir, sheet_name) / "review_rows.json"
    if f.exists():
        try:
            return json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            pass
    return []


def _mark_quick_sheet_complete(output_dir, sheet_name):
    cd = _quick_checkpoint_dir(output_dir, sheet_name)
    cd.mkdir(parents=True, exist_ok=True)
    atomic_write_text(cd / "_all_done", json.dumps(
        {"phase": "all", "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S")}, ensure_ascii=False))


def _is_quick_sheet_complete(output_dir, sheet_name) -> bool:
    return (_quick_checkpoint_dir(output_dir, sheet_name) / "_all_done").exists()


def _restore_enforce_backup(df, output_dir, sheet_name) -> int:
    """把重譯（enforce）備份套回 df，僅在重譯確實被中斷時還原（依 enforce_tag 過濾 part）。"""
    cp = Path(output_dir) / "_checkpoint" / _sanitize_sheet_name(sheet_name)
    ecp = cp / "enforce_checkpoint.json"
    if not ecp.exists():
        return 0
    try:
        enforce_tag = json.loads(ecp.read_text(encoding="utf-8")).get("enforce_tag")
    except Exception:
        enforce_tag = None
    if not enforce_tag:
        return 0
    backup = {}
    for f in sorted(cp.glob(f"part_{enforce_tag}_*.json")):
        try:
            backup.update(json.loads(f.read_text(encoding="utf-8")))
        except Exception:
            pass
    restore_count = 0
    for idx_str, trans in backup.items():
        idx = int(idx_str)
        if idx in df.index:
            df.at[idx, "translation"] = trans
            restore_count += 1
    return restore_count


async def _quick_phase4a(sheet_name, st, glossary, workplace_str, pool):
    df = st["df"]
    _timestamp(f"P4a 開始：{sheet_name}...")
    restored = _restore_enforce_backup(df, workplace_str, sheet_name)
    if restored > 0:
        print(f"    從重譯 checkpoint 還原 {restored} 條修正")
    df, review_df = await retry_protect(df, glossary, workplace_str, sheet_name=sheet_name, pool=pool)
    st["df"] = df
    st["review_df"] = review_df


def generate_quick_reports(excel_path, all_translated_dfs, all_review_rows, output_dir, overall_start):
    from openpyxl.styles import Color, PatternFill
    out_path = Path(output_dir) / f"{Path(excel_path).stem}_quick_proofread_output.xlsx"
    if all_translated_dfs:
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            for s_name, df_s in all_translated_dfs.items():
                df_s.to_excel(writer, index=False, sheet_name=s_name)
        print(f"\n  快速校對輸出：{out_path}")

    rp = Path(output_dir) / "quick_proofread_report.xlsx"
    if all_review_rows:
        combined = pd.concat(all_review_rows, ignore_index=True)
    else:
        combined = pd.DataFrame(columns=["sheet_name", "english", "current_translation", "category", "issue"])
    with pd.ExcelWriter(rp, engine="openpyxl") as writer:
        combined.to_excel(writer, index=False, sheet_name="Sheet1")
        ws = writer.sheets["Sheet1"]
        fill_glossary = PatternFill(start_color=Color(rgb="D9D9D9"), end_color=Color(rgb="D9D9D9"), fill_type="solid")
        fill_placeholder = PatternFill(start_color=Color(rgb="BDD7EE"), end_color=Color(rgb="BDD7EE"), fill_type="solid")
        fill_untranslated = PatternFill(start_color=Color(rgb="F4CCCC"), end_color=Color(rgb="F4CCCC"), fill_type="solid")
        fill_space = PatternFill(start_color=Color(rgb="E2EFDA"), end_color=Color(rgb="E2EFDA"), fill_type="solid")
        ic = None
        for ci, cell in enumerate(ws[1], start=1):
            if cell.value == "issue":
                ic = ci
                break
        if ic:
            for row in ws.iter_rows(min_row=2):
                v = str(row[ic - 1].value) if row[ic - 1].value else ""
                fill = (fill_glossary if v.startswith("[glossary]")
                        else fill_placeholder if v.startswith("[placeholder]")
                        else fill_untranslated if v.startswith("[untranslated]")
                        else fill_space if v.startswith("[space]") else None)
                if fill:
                    for cell in row:
                        cell.fill = fill
    print(f"  quick_proofread_report.xlsx 已寫入：{rp}（{len(combined)} 條）")

    qcp = _quick_checkpoint_dir(output_dir)
    if qcp.exists():
        import shutil
        shutil.rmtree(qcp)
        print(f"  已清除快速校對 checkpoint")
    ecp = Path(output_dir) / "_checkpoint"
    if ecp.exists():
        import shutil
        shutil.rmtree(ecp)
        print(f"  已清除強制檢查 checkpoint")
    elapsed(overall_start, "總計")


async def run_quick_proofread(excel_path, glossary_path, glossary_sheets, sheet_names, sheet_configs, session=None):
    overall_start = datetime.now()
    workplace = Path(excel_path).parent
    _timestamp("正在備份目標檔案...")
    backup_target_file(excel_path)
    _timestamp("正在載入術語庫...")
    t0 = datetime.now()
    if str(glossary_path) == "__AUTO__":
        glossary = auto_extract_glossary(excel_path, str(workplace))
    else:
        glossary = load_glossary(glossary_path, glossary_sheets)
    elapsed(t0, "術語庫")
    print(f"  術語庫載入完成：{len(glossary)} 條")

    all_translated_dfs = {}
    all_review_rows = []
    if session and session.get("accumulated_review_rows"):
        all_review_rows.append(pd.DataFrame(session["accumulated_review_rows"]))
        print(f"    還原 {len(session['accumulated_review_rows'])} 條審查記錄")
    completed_sheets = []
    pending_sheets = list(sheet_names)
    _save_quick_session(str(workplace), _build_session_data(
        excel_path, glossary_path, glossary_sheets,
        pending_sheets[0] if pending_sheets else "",
        [], pending_sheets, sheet_configs, mode="quick_proofread"))

    pool = SharedBatchPool()
    try:
        await pool.start()
    except RuntimeError:
        qcp = _quick_checkpoint_dir(str(workplace))
        if qcp.exists():
            import shutil
            shutil.rmtree(qcp)
        raise
    active = [c for c in pool.api_configs if not c.is_permanently_disabled]
    print(f"  已載入 {len(active)} 個可用 API：")
    for cfg in active:
        print(f"    {cfg.api_id} {cfg.model_provider}/{cfg.model} ({cfg.api_type}, 並發={cfg.parallel_limit})")

    try:
        sheet_states = {}
        for sheet_name in pending_sheets:
            df_original = pd.read_excel(excel_path, sheet_name=sheet_name, dtype=str)
            print(f"    總行數：{len(df_original)}")
            sheet_cfg = sheet_configs.get(sheet_name, {})
            selected_indices = set(sheet_cfg.get("selected_indices", []))
            if _is_quick_sheet_complete(str(workplace), sheet_name):
                df_full = df_original.copy()
                saved = _load_quick_sheet_translations(str(workplace), sheet_name)
                restore_count = 0
                for idx_str, trans in saved.items():
                    idx = int(idx_str)
                    if idx in df_full.index:
                        df_full.at[idx, "translation"] = trans
                        restore_count += 1
                if restore_count > 0:
                    print(f"    從 checkpoint 還原 {restore_count} 條修正")
                rows = _load_quick_sheet_review(str(workplace), sheet_name)
                if rows:
                    all_review_rows.append(pd.DataFrame(rows))
                    print(f"    還原 {len(rows)} 條審查記錄")
                all_translated_dfs[sheet_name] = df_full
                completed_sheets.append(sheet_name)
                print(f"  已從 checkpoint 還原 {sheet_name}")
                continue
            if not selected_indices:
                print(f"  {sheet_name} 未選取任何條目，跳過")
                all_translated_dfs[sheet_name] = df_original
                completed_sheets.append(sheet_name)
                continue
            df = df_original[df_original.index.isin(selected_indices)].copy()
            sheet_states[sheet_name] = {"df_original": df_original, "df": df}

        t = datetime.now()
        _timestamp("Phase 4a：重譯修正（全部工作表並行）...")
        results = await asyncio.gather(*[
            _quick_phase4a(sheet_name, st, glossary, str(workplace), pool)
            for sheet_name, st in sheet_states.items()
        ], return_exceptions=True)
        for sheet_name, res in zip(sheet_states, results):
            if isinstance(res, BaseException):
                if isinstance(res, RuntimeError) and "all APIs permanently disabled" in str(res):
                    print("\n  ⚠️所有 API 已永久停用，快速校對中止")
                raise res
        elapsed(t, "P4a")

        for sheet_name, st in sheet_states.items():
            df = st["df"]
            review_df = st["review_df"]
            if review_df is not None and not review_df.empty:
                review_df.insert(0, "sheet_name", sheet_name)
                all_review_rows.append(review_df)
            ecp = Path(str(workplace)) / "_checkpoint" / _sanitize_sheet_name(sheet_name) / "enforce_checkpoint.json"
            if ecp.exists():
                ecp.unlink()
            st["df_original"].update(df)
            all_translated_dfs[sheet_name] = st["df_original"]
            saved = {}
            for idx in df.index:
                trans = df.at[idx, "translation"]
                if trans is not None:
                    saved[str(idx)] = str(trans)
            _save_quick_sheet_translations(str(workplace), sheet_name, saved)
            if review_df is not None and not review_df.empty:
                _save_quick_sheet_review(str(workplace), sheet_name, review_df.to_dict(orient="records"))
            _mark_quick_sheet_complete(str(workplace), sheet_name)
            completed_sheets.append(sheet_name)
            print(f"  {sheet_name} 完成")

        pending_sheets.clear()
        _save_quick_session(str(workplace), _build_session_data(
            excel_path, glossary_path, glossary_sheets,
            "", completed_sheets, pending_sheets, sheet_configs, mode="quick_proofread"))
    finally:
        await pool.close()

    # Phase 4b：模板校正（在所有工作表完成後執行；TEMPLATES 為空時無作用）
    t_p4b = datetime.now()
    _timestamp("Phase 4b：模板校正...")
    for sheet_name in all_translated_dfs:
        sheet_cfg = sheet_configs.get(sheet_name, {})
        selected_indices = set(sheet_cfg.get("selected_indices", []))
        if not selected_indices:
            continue
        df_original = all_translated_dfs[sheet_name]
        df = df_original[df_original.index.isin(selected_indices)].copy()
        df = template_correction(df, glossary, str(workplace), sheet_name=sheet_name)
        df_original.update(df)
        all_translated_dfs[sheet_name] = df_original
    elapsed(t_p4b, "P4b")

    generate_quick_reports(excel_path, all_translated_dfs, all_review_rows, str(workplace), overall_start)
    _remove_session_file(str(workplace))

def main():
    print("校對模組已載入。請使用 batch_proofread.py 執行 CLI。")

if __name__ == "__main__":
    main()
