#!/usr/bin/env python3
import sys
import os
import json
from pathlib import Path
from datetime import datetime
import pandas as pd

sys.path.insert(0, str(Path(__file__).parent))

from glossary import load_glossary
from tm_matcher import match_and_fill
from llm_translator import translate_all as llm_translate
from llm_translator import save_session, delete_checkpoint_files, load_backup
from enforcer import enforce


SEP = "=" * 60
AUTO_GLOSSARY_SENTINEL = Path("__AUTO__")
WORKPLACE_DIR_NAME = "workplace"


def _get_workplace() -> Path:
    """取得腳本所在目錄同級的 workplace/ 目錄，不存在則自動建立。"""
    wp = Path(__file__).parent / WORKPLACE_DIR_NAME
    wp.mkdir(parents=True, exist_ok=True)
    return wp


def timestamp(msg: str):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"  [{now}] {msg}")


def elapsed(start: datetime, label: str):
    secs = (datetime.now() - start).total_seconds()
    print(f"  [{label}] 耗時: {secs:.1f} 秒")


def list_excel_files(directory: Path) -> list[Path]:
    if not directory.exists():
        return []
    return sorted([f for f in directory.glob("*.xlsx") if not f.name.startswith("~$")])


def choose_int(prompt: str, max_val: int, allow_zero: bool = False) -> int | None:
    while True:
        try:
            choice = input(prompt).strip()
            if not choice:
                return None
            idx = int(choice)
            if idx == 0 and allow_zero:
                return 0
            if 1 <= idx <= max_val:
                return idx
        except ValueError:
            pass
        print("  無效輸入，請重新輸入。")


def choose_from_list(items: list[str], title: str, allow_zero: bool = False) -> int | None:
    print(f"\n{SEP}")
    print(f"  {title}")
    print(SEP)
    for i, item in enumerate(items, 1):
        print(f"  [{i}] {item}")
    if allow_zero:
        print("  [0] 不使用 / 跳過")
    return choose_int("\n請輸入編號: ", len(items), allow_zero=allow_zero)


def choose_multi(items: list[str], title: str) -> list[str]:
    print(f"\n{SEP}")
    print(f"  {title}")
    print(SEP)
    for i, item in enumerate(items, 1):
        print(f"  [{i}] {item}")
    print("  [0] 全部選取")
    while True:
        try:
            choice = input("\n請輸入編號（可多個以逗號分隔）: ").strip()
            if choice == "0":
                return list(items)
            indices = [int(x.strip()) for x in choice.split(",") if x.strip()]
            selected = [items[i-1] for i in indices if 1 <= i <= len(items)]
            if selected:
                return selected
        except (ValueError, IndexError):
            pass
        print("  無效輸入，請重新輸入。")


def step1() -> Path:
    """step1：從 workplace/ 選擇翻譯目標 Excel，並輸入語言代碼。"""
    workplace = _get_workplace()
    xlsx_files = list_excel_files(workplace)
    if not xlsx_files:
        print(f"\n錯誤：{workplace}/ 下沒有 Excel 檔案")
        print(f"請將翻譯目標 Excel 放入 {workplace}/ 後再執行。")
        input("\n按 Enter 關閉...")
        sys.exit(1)
    fnames = [f.name for f in xlsx_files]
    c = choose_from_list(fnames, f"請選擇目標 Excel 檔案（{workplace}/）：")
    if c is None:
        print("已取消。")
        input("\n按 Enter 關閉...")
        sys.exit(0)
    return xlsx_files[c - 1]


def step2(exclude_path: Path | None = None) -> tuple[Path | None, list[str] | None]:
    """step2：從 workplace/ 選擇術語庫 Excel。"""
    workplace = _get_workplace()
    xlsx_files = [f for f in list_excel_files(workplace)
                  if exclude_path is None or f.resolve() != exclude_path.resolve()]
    if not xlsx_files:
        inp = input(f"\n{workplace}/ 下沒有 Excel 檔案，輸入其他路徑或直接 Enter 跳過: ").strip()
        if inp:
            p = Path(inp)
            if p.exists():
                return p, None
        return None, None
    fnames = [f.name for f in xlsx_files]
    fnames.append("✨ 自動從目標 Excel 萃取（name/manual 工作表）")
    c = choose_from_list(fnames, "選擇術語庫 Excel（translation 有值的行即為術語）:", allow_zero=True)
    if c is None or c == 0:
        return None, None
    if c == len(fnames):
        return AUTO_GLOSSARY_SENTINEL, None
    gp = xlsx_files[c - 1]
    xls = pd.ExcelFile(gp)
    sheets = xls.sheet_names
    if len(sheets) == 1:
        print(f"  自動使用工作表: {sheets[0]}")
        return gp, sheets
    selected = choose_multi(sheets, "選擇哪些工作表作為術語來源：")
    return gp, selected


def step3(excel_path: Path) -> list[str]:
    """step3：選擇翻譯目標工作表。"""
    xls = pd.ExcelFile(excel_path)
    sheets = xls.sheet_names
    if len(sheets) == 1:
        return sheets
    return choose_multi(sheets, "選擇翻譯目標工作表：")

def step4_choose_sheet_mode(sheet_name: str, untranslated_count: int) -> dict:
    """
    選擇單一工作表的翻譯範圍模式。
    回傳模式 dict:
      {"type": "all"}
      {"type": "first_n", "n": N}
      {"type": "range", "start": S, "end": E}
    """
    print(f"  未翻譯: {untranslated_count} 條")
    print(f"  [1] 全部未翻譯條目")
    print(f"  [2] 僅測試前 N 條")
    print(f"  [3] 指定行數範圍")
    while True:
        c = input("\n請選擇: ").strip()
        if c == "1":
            print(f"  選取全部 {untranslated_count} 條未翻譯")
            return {"type": "all"}
        elif c == "2":
            try:
                n = int(input("  請輸入 N: ").strip())
                print(f"  選取前 {n} 條未翻譯")
                return {"type": "first_n", "n": n}
            except ValueError:
                print("  無效數字")
        elif c == "3":
            try:
                rng = input("  請輸入行數範圍（如 100-500）: ").strip()
                parts = rng.split("-")
                s, e = int(parts[0]) - 1, int(parts[1])
                print(f"  選取行 {s+1} 到 {e}")
                return {"type": "range", "start": s, "end": e}
            except (ValueError, IndexError):
                print("  無效範圍")
        else:
            print("  無效輸入")

def step4_apply_mode(df_full: pd.DataFrame, mode: dict) -> pd.DataFrame:
    """根據預選模式對特定工作表的 DataFrame 套用範圍。"""
    untranslated_mask = df_full["translation"].isna() | (df_full["translation"].isnull())
    if mode["type"] == "all":
        return df_full[untranslated_mask].copy()
    elif mode["type"] == "first_n":
        idx = list(df_full.index[untranslated_mask][:mode["n"]])
        return df_full.loc[idx].copy() if idx else df_full.iloc[:0].copy()
    elif mode["type"] == "range":
        return df_full.iloc[mode["start"]:mode["end"]].copy()
    return df_full[untranslated_mask].copy()


def confirm(excel_path: Path, glossary_path: Path | None,
            sheets: list[str], total_rows: int) -> bool:
    """顯示翻譯摘要並要求確認。"""
    print(f"\n{SEP}")
    print("  即將執行以下操作：")
    print(SEP)
    print(f"  目標: {excel_path}")
    if glossary_path == AUTO_GLOSSARY_SENTINEL:
        print(f"  術語庫: 自動萃取（name/manual 工作表）")
    else:
        print(f"  術語庫: {glossary_path if glossary_path else '無（僅內建 ADD_LIST）'}")
    print(f"  工作表: {', '.join(sheets)}")
    print(f"  翻譯條數: {total_rows}")
    print(SEP)
    return input("\n確認執行？(Y/n): ").strip().lower() != "n"


def find_existing_progress() -> dict | None:
    """掃描 workplace/_checkpoint/session.json，偵測未完成的翻譯進度。"""
    workplace = _get_workplace()
    sf = workplace / "_checkpoint" / "session.json"
    if sf.exists():
        try:
            session = json.loads(sf.read_text(encoding="utf-8"))
            session["_workplace"] = str(workplace)
            return session
        except Exception:
            pass
    return None

def prompt_resume(session: dict) -> bool:
    """顯示續傳提示（支援新舊格式）。"""
    print(f"\n{SEP}")
    print("  偵測到未完成的翻譯進度！")
    print(SEP)
    excel_name = Path(session.get("excel_path", "?")).name
    print(f"  Excel: {excel_name}")

    if "current_sheet" in session:
        # 新版多工作表格式
        cs = session["current_sheet"]
        done = len(session.get("completed_sheets", []))
        total = len(session.get("pending_sheets", [])) + done
        print(f"  工作進度: {done}/{total} 工作表")
        print(f"  當前工作表: {cs}")
    else:
        # 舊版單工作表格式
        sn = session.get("sheet_name", "?")
        total = session.get("total_selected", "?")
        print(f"  工作表: {sn}")
        print(f"  條數: {total}")

    glossary_path = session.get("glossary_path")
    if glossary_path == "__AUTO__":
        print(f"  術語庫: 自動萃取（name/manual 工作表）")
    else:
        print(f"  術語庫: {Path(glossary_path).name if glossary_path else '無'}")
    print(SEP)
    return input("\n繼續翻譯？(Y/n): ").strip().lower() != "n"

def _build_session_data(excel_path, glossary_path, glossary_sheets,
                        current_sheet, completed_sheets, pending_sheets, sheet_configs,
                        accumulated_review_rows=None):
    data = {
        "excel_path": str(excel_path),
        "glossary_path": str(glossary_path) if glossary_path else None,
        "glossary_sheets": glossary_sheets,
        "current_sheet": current_sheet,
        "completed_sheets": list(completed_sheets),
        "pending_sheets": list(pending_sheets),
        "sheet_configs": sheet_configs,
        "timestamp": datetime.now().isoformat(),
    }
    if accumulated_review_rows:
        # 扁平化所有 DataFrame 為 dict 列表
        rows = []
        for df in accumulated_review_rows:
            for _, r in df.iterrows():
                rows.append(r.to_dict())
        data["accumulated_review_rows"] = rows
    return data

def _write_outputs(excel_path: Path, workplace: Path, all_translated_dfs: dict,
                   all_review_rows: list[pd.DataFrame], overall_start: datetime):
    out_path = workplace / f"{excel_path.stem}_translated_output.xlsx"
    timestamp(f"寫入輸出檔案 {out_path.name}...")
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        for sheet_name, df_data in all_translated_dfs.items():
            df_data.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"  輸出完成: {out_path}")

    if all_review_rows:
        combined_review = pd.concat(all_review_rows, ignore_index=True)
        report_path = workplace / "review_report.xlsx"
        combined_review.to_excel(report_path, index=False, sheet_name='Sheet1')

        # ── 補上顏色標記（與 enforcer.py 一致） ──
        from openpyxl.styles import PatternFill
        from openpyxl import load_workbook

        wb = load_workbook(report_path)
        ws = wb['Sheet1']

        issue_col = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == "issue":
                issue_col = col_idx
                break

        if issue_col:
            fill_glossary = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            fill_placeholder = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            fill_untranslated = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

            for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
                issue_val = str(row[issue_col - 1].value) if row[issue_col - 1].value else ""
                if issue_val.startswith("[glossary]"):
                    fill = fill_glossary
                elif issue_val.startswith("[placeholder]"):
                    fill = fill_placeholder
                elif issue_val.startswith("[untranslated]"):
                    fill = fill_untranslated
                else:
                    continue
                for cell in row:
                    cell.fill = fill

        wb.save(report_path)
        print(f"  審查報告已寫入: {report_path} ({len(combined_review)} 條)")

    delete_checkpoint_files(str(workplace))
    print(f"  已清除進度檔案")

    total_secs = (datetime.now() - overall_start).total_seconds()
    print(f"\n{SEP}")
    timestamp("管線執行完成")
    print(f"  總耗時: {total_secs:.1f} 秒")
    print(SEP)
    input("\n按 Enter 關閉...")

def main():
    print(SEP)
    print("      混合翻譯管線 v1.5")
    print(f"      工作目錄: {_get_workplace()}")
    print(SEP)

    # ── 檢查上次的 debug 訊息 ──
    debug_dir = _get_workplace() / "_debugmessage"
    if debug_dir.exists():
        debug_files = sorted(debug_dir.glob("debug_*.json"))
        if debug_files:
            print(f"\n{'=' * 60}")
            print(f"  ⚠️ 上次執行有低翻譯率記錄")
            print(f"{'=' * 60}")
            for f in debug_files:
                try:
                    data = json.loads(f.read_text(encoding="utf-8"))
                    print(f"  ─ [{data.get('api_id', '?')}] {data.get('model', '?')} ─")
                    print(f"    批次大小: {data.get('batch_size')}, "
                          f"成功: {data.get('success')}/{data.get('returned')}, "
                          f"翻譯率: {data.get('rate', 0):.1%}")
                    preview = data.get("response_preview", "")
                    if preview:
                        print(f"    回傳內容:\n{preview}")
                except Exception as e:
                    print(f"  無法讀取 {f.name}: {e}")
            print(f"{'=' * 60}\n")
        import shutil
        shutil.rmtree(debug_dir)
        print(f"  已清除 debug 訊息\n")

    # ── 續傳檢測 ──
    exist_session = find_existing_progress()
    if exist_session:
        if not prompt_resume(exist_session):
            workplace = _get_workplace()
            delete_checkpoint_files(str(workplace))
            print("\n  已清除舊進度，開始新的翻譯。\n")
        else:
            # 解析 session（相容新舊格式）
            excel_path = Path(exist_session["excel_path"])
            glossary_path_str = exist_session.get("glossary_path")
            glossary_path = Path(glossary_path_str) if glossary_path_str else None
            glossary_sheets = exist_session.get("glossary_sheets")
            workplace = _get_workplace()

            if "current_sheet" in exist_session:
                # 新版多工作表格式
                pending_sheets = list(exist_session.get("pending_sheets", []))
                completed_sheets = list(exist_session.get("completed_sheets", []))
                sheet_configs = exist_session.get("sheet_configs", {})
            else:
                # 舊版單工作表格式 → 轉換為新版
                sn = exist_session.get("sheet_name", "")
                sel = list(exist_session.get("selected_indices", []))
                pending_sheets = [sn] if sn else []
                completed_sheets = []
                sheet_configs = {sn: {"selected_indices": sel}} if sn else {}

            # 載入術語庫
            overall_start = datetime.now()
            t0 = datetime.now()
            timestamp("載入術語庫...")
            if glossary_path_str == "__AUTO__":
                from glossary import auto_extract_glossary
                glossary = auto_extract_glossary(excel_path, workplace)
            else:
                glossary = load_glossary(glossary_path, glossary_sheets if glossary_path else None)
            elapsed(t0, "載入術語庫")
            print(f"  術語庫載入完成: {len(glossary)} 條")

            all_translated_dfs = {}
            all_review_rows = []

            # ── 還原已完成工作表資料 ──
            from llm_translator import load_backup
            for sn in completed_sheets:
                timestamp(f"從檢查點還原工作表: {sn}...")
                df_full = pd.read_excel(excel_path, sheet_name=sn, dtype=str)
                backup = load_backup(str(workplace), sn)
                if backup:
                    restore_count = 0
                    for idx_str, trans in backup.items():
                        idx = int(idx_str)
                        if idx in df_full.index:
                            df_full.at[idx, "translation"] = trans
                            restore_count += 1
                    print(f"    還原 {restore_count} 條翻譯")
                all_translated_dfs[sn] = df_full

            # ── 還原累積審查資料 ──
            if "accumulated_review_rows" in exist_session:
                rows = exist_session["accumulated_review_rows"]
                if rows:
                    all_review_rows.append(pd.DataFrame(rows))
                    print(f"    還原 {len(rows)} 條審查記錄")

            while pending_sheets:
                sheet_name = pending_sheets[0]

                print(f"\n{SEP}")
                print(f"  處理工作表: {sheet_name}")
                print(SEP)

                t1 = datetime.now()
                timestamp(f"讀取 {sheet_name}...")
                df_full = pd.read_excel(excel_path, sheet_name=sheet_name, dtype=str)
                print(f"    總條數: {len(df_full)}")

                sheet_cfg = sheet_configs.get(sheet_name, {})
                selected_indices = set(sheet_cfg.get("selected_indices", []))
                if selected_indices:
                    mask = df_full.index.isin(selected_indices) & (
                        df_full["translation"].isna() | (df_full["translation"].isnull())
                    )
                    df = df_full[mask].copy()
                    print(f"    還原選取範圍，剩餘未翻譯: {len(df)} 條")
                else:
                    df = df_full[df_full["translation"].isna() | (df_full["translation"].isnull())].copy()
                    print(f"    無法還原選取範圍（相容模式），未翻譯: {len(df)} 條")

                if len(df) == 0:
                    print(f"  {sheet_name} 無需翻譯，跳過")
                    all_translated_dfs[sheet_name] = df_full
                    pending_sheets.pop(0)
                    completed_sheets.append(sheet_name)
                    save_session(str(workplace), _build_session_data(
                        excel_path, glossary_path, glossary_sheets,
                        pending_sheets[0] if pending_sheets else "",
                        completed_sheets, pending_sheets, sheet_configs,
                        accumulated_review_rows=all_review_rows))
                    continue

                t2 = datetime.now()
                timestamp("模板參數化比對...")
                df = match_and_fill(df, glossary)
                matched = int((df["_status"] == "已處理").sum())
                elapsed(t2, "模板比對")
                print(f"    模板匹配完成: {matched} 條已處理")

                pending_count = int((df["translation"].isna() | (df["translation"] == "nan")).sum())
                timestamp(f"LLM 批次翻譯（待翻譯: {pending_count} 條）...")
                t3 = datetime.now()
                if pending_count > 0:
                    df = llm_translate(df, glossary, str(workplace), sheet_name=sheet_name)
                else:
                    print("    無需 LLM 翻譯")
                elapsed(t3, "LLM 翻譯")

                t4 = datetime.now()
                timestamp("術語強制後處理...")
                df, review_df = enforce(df, glossary, str(workplace),
                                        report_name=f"review_report_{sheet_name}.xlsx",
                                        write_report=False, sheet_name=sheet_name)
                if not review_df.empty:
                    review_df.insert(0, "sheet_name", sheet_name)
                    all_review_rows.append(review_df)
                elapsed(t4, "術語強制後處理")

                df_full.update(df)
                all_translated_dfs[sheet_name] = df_full

                pending_sheets.pop(0)
                completed_sheets.append(sheet_name)
                next_sheet = pending_sheets[0] if pending_sheets else ""
                save_session(str(workplace), _build_session_data(
                    excel_path, glossary_path, glossary_sheets,
                    next_sheet, completed_sheets, pending_sheets, sheet_configs,
                    accumulated_review_rows=all_review_rows))
                print(f"  ✅ {sheet_name} 完成")

            _write_outputs(excel_path, workplace, all_translated_dfs, all_review_rows, overall_start)
            return

    # ── 全新翻譯路徑 ──
    excel_path = step1()
    workplace = _get_workplace()
    glossary_path, glossary_sheets = step2(exclude_path=excel_path)
    sheet_names = step3(excel_path)

    overall_start = datetime.now()
    t0 = datetime.now()
    timestamp("載入術語庫...")
    if glossary_path == AUTO_GLOSSARY_SENTINEL:
        from glossary import auto_extract_glossary
        glossary = auto_extract_glossary(excel_path, workplace)
    else:
        glossary = load_glossary(glossary_path, glossary_sheets)
    elapsed(t0, "載入術語庫")
    print(f"  術語庫載入完成: {len(glossary)} 條")

    # 收集各工作表的未翻譯統計（同時快取 DataFrame，減少重複讀取）
    sheet_untranslated = []
    cached_dfs = {}
    for sn in sheet_names:
        df_temp = pd.read_excel(excel_path, sheet_name=sn, dtype=str)
        uc = int(df_temp["translation"].isna().sum()) if "translation" in df_temp.columns else len(df_temp)
        sheet_untranslated.append((sn, uc))
        cached_dfs[sn] = df_temp

    # 逐工作表選擇翻譯範圍
    sheet_configs = {}
    for sn, uc in sheet_untranslated:
        print(f"\n{SEP}")
        print(f"  設定工作表：{sn}（{uc} 條未翻譯）")
        print(SEP)
        sn_mode = step4_choose_sheet_mode(sn, uc)
        df_filtered = step4_apply_mode(cached_dfs[sn], sn_mode)
        sheet_configs[sn] = {
            "mode": sn_mode,
            "selected_indices": list(df_filtered.index) if len(df_filtered) > 0 else [],
        }
    del cached_dfs  # 釋放記憶體

    # 顯示摘要並確認
    total_selected = sum(len(cfg["selected_indices"]) for cfg in sheet_configs.values())
    if total_selected == 0:
        print("\n所有工作表都沒有符合條件的條目，無需處理")
        input("\n按 Enter 關閉...")
        return
    if not confirm(excel_path, glossary_path, sheet_names, total_selected):
        print("  已取消。")
        input("\n按 Enter 關閉...")
        return

    # 儲存初始 session（多工作表格式）
    pending_sheets = list(sheet_names)
    save_session(str(workplace), _build_session_data(
        excel_path, glossary_path, glossary_sheets,
        pending_sheets[0] if pending_sheets else "",
        [], pending_sheets, sheet_configs))

    # ── 逐一處理所有工作表 ──
    all_translated_dfs = {}
    all_review_rows = []
    completed_sheets = []

    while pending_sheets:
        sheet_name = pending_sheets[0]
        print(f"\n{SEP}")
        print(f"  處理工作表: {sheet_name}")
        print(SEP)

        t1 = datetime.now()
        timestamp(f"讀取 {sheet_name}...")
        df_full = pd.read_excel(excel_path, sheet_name=sheet_name, dtype=str)
        print(f"    總條數: {len(df_full)}")

        sheet_cfg = sheet_configs.get(sheet_name, {})
        selected_indices = set(sheet_cfg.get("selected_indices", []))
        if not selected_indices:
            print(f"  {sheet_name} 沒有選取任何條目，跳過")
            all_translated_dfs[sheet_name] = df_full
            pending_sheets.pop(0)
            completed_sheets.append(sheet_name)
            next_sheet = pending_sheets[0] if pending_sheets else ""
            save_session(str(workplace), _build_session_data(
                excel_path, glossary_path, glossary_sheets,
                next_sheet, completed_sheets, pending_sheets, sheet_configs,
                accumulated_review_rows=all_review_rows))
            continue

        mask = df_full.index.isin(selected_indices) & (
            df_full["translation"].isna() | (df_full["translation"].isnull())
        )
        df = df_full[mask].copy()
        remaining = len(df)
        print(f"    還原選取範圍，剩餘未翻譯: {remaining} 條")

        if remaining == 0:
            print(f"  {sheet_name} 已全部翻譯完成，跳過")
            all_translated_dfs[sheet_name] = df_full
            pending_sheets.pop(0)
            completed_sheets.append(sheet_name)
            next_sheet = pending_sheets[0] if pending_sheets else ""
            save_session(str(workplace), _build_session_data(
                excel_path, glossary_path, glossary_sheets,
                next_sheet, completed_sheets, pending_sheets, sheet_configs,
                accumulated_review_rows=all_review_rows))
            continue

        t2 = datetime.now()
        timestamp("模板參數化比對...")
        df = match_and_fill(df, glossary)
        matched = int((df["_status"] == "已處理").sum())
        elapsed(t2, "模板比對")
        print(f"    模板匹配完成: {matched} 條已處理")

        pending_count = int((df["translation"].isna() | (df["translation"] == "nan")).sum())
        timestamp(f"LLM 批次翻譯（待翻譯: {pending_count} 條）...")
        t3 = datetime.now()
        if pending_count > 0:
            df = llm_translate(df, glossary, str(workplace), sheet_name=sheet_name)
        else:
            print("    無需 LLM 翻譯")
        elapsed(t3, "LLM 翻譯")

        t4 = datetime.now()
        timestamp("術語強制後處理...")
        df, review_df = enforce(df, glossary, str(workplace),
                                report_name=f"review_report_{sheet_name}.xlsx",
                                write_report=False, sheet_name=sheet_name)
        if not review_df.empty:
            review_df.insert(0, "sheet_name", sheet_name)
            all_review_rows.append(review_df)
        elapsed(t4, "術語強制後處理")

        df_full.update(df)
        all_translated_dfs[sheet_name] = df_full

        # 更新進度
        pending_sheets.pop(0)
        completed_sheets.append(sheet_name)
        next_sheet = pending_sheets[0] if pending_sheets else ""
        save_session(str(workplace), _build_session_data(
            excel_path, glossary_path, glossary_sheets,
            next_sheet, completed_sheets, pending_sheets, sheet_configs,
            accumulated_review_rows=all_review_rows))
        print(f"  ✅ {sheet_name} 完成")

    # ── 所有工作表完成：寫入輸出 ──
    _write_outputs(excel_path, workplace, all_translated_dfs, all_review_rows, overall_start)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n發生錯誤: {e}")
        import traceback
        traceback.print_exc()
        input("\n按 Enter 關閉...")