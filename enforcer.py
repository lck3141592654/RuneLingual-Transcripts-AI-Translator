import pandas as pd
from pathlib import Path
import json, asyncio, time, re
from dotenv import load_dotenv
from datetime import datetime

load_dotenv()

from glossary import normalize_term
from llm_translator import BATCH_SIZE_LIMIT, atomic_write_text
from api_config import API_TIMEOUT, OVER_RETURN_TOLERANCE
from shared_pool import SharedBatchPool

RETRY_PROMPTS = [
    "1. 禁止翻譯 [] 内的内容，如果译文中 [] 内的内容与原文不同，必須恢复为原文的占位符。"
    "2. 如果当前译文与原文完全相同或沒有內容，请重新翻译。"
    "3. 使用简体中文修正以下翻译中的游戏术语错误。"
    "4. 如果术语表的人名、地名等名詞用'.'分隔，例如'索菲娅.休斯'，你必須使用'.'分隔，不要自行改為'·'或其他方式：",
    "术语仍不正确，或占位符被翻译，或译文与原文相同、沒有內容，请严格对照要求重新翻译。"
    "如果术语表的人名、地名等名詞用'.'分隔，例如'索菲娅.休斯'，你必須使用'.'分隔，不要自行改為'·'或其他方式：",
    "最后一次修正，以下条目的术语必须使用指定翻译，占位符必须保留原样，且不能回吐原文。"
    "如果术语表的人名、地名等名詞用'.'分隔，例如'索菲娅.休斯'，你必須使用'.'分隔，不要自行改為'·'或其他方式：",
]

_re_cache: dict = {}


def _cached_patterns(term: str):
    """Cache compiled regex per glossary term; returns (pat, norm_pat)."""
    cached = _re_cache.get(term)
    if cached is None:
        pat = re.compile(r"(?<![a-z'])" + re.escape(term.lower()) + r"(?![a-z'])")
        norm = normalize_term(term)
        norm_pat = None
        if norm != term.lower():
            norm_pat = re.compile(r"(?<![a-z'])" + re.escape(norm) + r"(?![a-z'])")
        cached = (pat, norm_pat)
        _re_cache[term] = cached
    return cached


def _find_matched_spans(eng_lower: str, glossary: dict) -> list:
    """對每條原文找出所有匹配到（詞邊界）的術語及其 span，按術語長度降序排列。

    回傳 list[(eng, chn, [(start, end), ...])]
    """
    matches = []
    for eng, chn in glossary.items():
        pat, norm_pat = _cached_patterns(eng)
        spans = [m.span() for m in pat.finditer(eng_lower)]
        if norm_pat is not None:
            spans.extend([m.span() for m in norm_pat.finditer(eng_lower)])
        if spans:
            matches.append((eng, chn, spans))
    # 長術語優先處理（避免短術語被長術語覆蓋時誤報）
    matches.sort(key=lambda x: len(x[0]), reverse=True)
    return matches



def check_glossary_usage(english_text: str, translated_text: str, glossary: dict) -> list:
    """詞邊界匹配，防止『V』匹配到『Convert』或『Don』匹配到『don't』這類誤報。

    當原文中長術語（如 "Lord Hosidius"）的翻譯已正確出現在譯文中，
    則其範圍內的短術語（如 "Hosidius"）不再重複檢查。
    """
    issues = []
    eng_lower = english_text.lower()
    trans_lower = translated_text.lower()

    matches = _find_matched_spans(eng_lower, glossary)

    covered_by_correct = []  # 已被正確翻譯的長術語所覆蓋的 span 集合
    for eng, chn, spans in matches:
        # 檢查此術語的所有匹配 span 是否都被已覆蓋
        all_covered = all(
            any(s >= cs[0] and e <= cs[1] for cs in covered_by_correct)
            for s, e in spans
        )
        if all_covered:
            continue  # 已被更長且翻譯正確的術語覆蓋，跳過

        if chn.lower() in trans_lower:
            # 翻譯正確，將此術語的 span 加入已覆蓋集合
            covered_by_correct.extend(spans)
        else:
            issues.append((eng, chn, f"术语「{eng}」应为「{chn}」"))

    return issues

def check_placeholder(english_text: str, translated_text: str) -> list[str]:
    """比對原文和譯文中的 [...] 是否一致，返回問題描述清單"""
    issues = []
    eng_placeholders = re.findall(r'\[.*?\]', english_text)
    for ph in eng_placeholders:
        if ph not in translated_text:
            issues.append(f"佔位符「{ph}」被翻譯或遺失")
    return issues


def check_untranslated(english_text: str, translated_text) -> list[str]:
    """檢查是否未成功翻譯，返回問題描述清單"""
    issues = []
    # 情況 A：translation 為空
    if translated_text is None or (isinstance(translated_text, float) and pd.isna(translated_text)):
        issues.append("translation 欄位為空")
        return issues
    trans_str = str(translated_text).strip()
    if trans_str.lower() in ("nan", "nat", "none"):
        issues.append("translation 欄位為空")
        return issues
    # 情況 B：譯文與原文相同（不分大小寫），且原文包含英文字母
    if trans_str.lower() == english_text.strip().lower():
        if re.search(r'[a-zA-Z]', english_text):
            # 排除純佔位符條目：去除 [] 後只留下空白或標點符號時不視為未翻譯
            stripped = re.sub(r'\[.*?\]', '', english_text).strip()
            if stripped and re.search(r'[a-zA-Z]', stripped):
                issues.append("譯文與原文相同，未實際翻譯")
    return issues


def check_space_issues(translated_text: str) -> list:
    """檢查譯文中的四種空格問題，回傳問題清單 list[(type, suggested_fix, desc)]"""
    issues = []
    if translated_text is None or (isinstance(translated_text, float) and pd.isna(translated_text)):
        return issues
    trans_str = str(translated_text)
    if '  ' in trans_str:
        issues.append(("space", "", "存在連續多個空格"))
    if re.search(r'[\u4e00-\u9fff] [\u4e00-\u9fff]', trans_str):
        issues.append(("space", "", "中文字間不應有空格"))
    if re.search(r' [\u4e00-\u9fff\uff0c\u3002\uff01\uff1f\uff1b\uff1a\u300c\u300d\u300e\u300f\uff08\uff09\u3010\u3011\u300a\u300b\u2014\u2014\u2026\u2026\u00b7\u3001]', trans_str):
        issues.append(("space", "", "空格後不應直接接中文標點符號"))
    if re.search(r'[\u4e00-\u9fff\uff0c\u3002\uff01\uff1f\uff1b\uff1a\u300c\u300d\u300e\u300f\uff08\uff09\u3010\u3011\u300a\u300b\u2014\u2014\u2026\u2026\u00b7\u3001] ', trans_str):
        issues.append(("space", "", "中文標點符號後不應有空格"))
    if re.search(r' [\.!\?,;:]', trans_str):
        issues.append(("space", "", "空格後不應直接接英文標點符號"))
    if re.search(r'[\.!\?,;:] ', trans_str):
        issues.append(("space", "", "英文標點符號後不應有空格"))
    return issues


def fix_single_placeholder(english_text: str, translated_text: str) -> str:
    """若原文只有一個 [] 佔位符，直接用原文的 [] 內容覆蓋譯文的 [] 內容"""
    eng_phs = re.findall(r'\[(.*?)\]', english_text)
    if len(eng_phs) != 1:
        return translated_text
    target_ph = f"[{eng_phs[0]}]"
    trans_phs = re.findall(r'\[(.*?)\]', translated_text)
    if trans_phs:
        old_ph = f"[{trans_phs[0]}]"
        if old_ph != target_ph:
            return translated_text.replace(old_ph, target_ph)
    return translated_text


def fix_space_issues(text: str) -> str:
    """清除四種不合格空格，保留正常空格"""
    text = re.sub(r' {2,}', ' ', text)
    text = re.sub(r'([\u4e00-\u9fff]) ([\u4e00-\u9fff])', r'\1\2', text)
    text = re.sub(r' ([\u4e00-\u9fff\uff0c\u3002\uff01\uff1f\uff1b\uff1a\u300c\u300d\u300e\u300f\uff08\uff09\u3010\u3011\u300a\u300b\u2014\u2014\u2026\u2026\u00b7\u3001])', r'\1', text)
    text = re.sub(r'([\u4e00-\u9fff\uff0c\u3002\uff01\uff1f\uff1b\uff1a\u300c\u300d\u300e\u300f\uff08\uff09\u3010\u3011\u300a\u300b\u2014\u2014\u2026\u2026\u00b7\u3001]) ', r'\1', text)
    text = re.sub(r' [\.!\?,;:]', r'', text)
    text = re.sub(r'[\.!\?,;:] ', r'', text)
    return text
def preprocess_issues(df, pool):
    """對問題條目進行腳本預處理（單佔位符修正 + 空格修正），直接修改 df"""
    for idx, row, issues in pool:
        eng = str(row.get("english", ""))
        trans = row.get("translation")
        if trans is None or (isinstance(trans, float) and pd.isna(trans)):
            continue
        new_trans = str(trans)
        new_trans = fix_single_placeholder(eng, new_trans)
        new_trans = fix_space_issues(new_trans)
        if new_trans != str(trans):
            df.at[idx, "translation"] = new_trans
    return df
def scan_issues(df, relevant_glossary) -> list[tuple]:
    """走訪所有行，彙整三種檢查的結果，回傳問題 pool"""
    pool = []
    for idx, row in df.iterrows():
        eng = str(row.get("english", ""))
        trans = row.get("translation")
        all_issues = []

        # 未翻譯檢查（包含空值情況）
        untrans_issues = check_untranslated(eng, trans)
        if untrans_issues:
            for ui in untrans_issues:
                all_issues.append(("untranslated", "", ui))
            pool.append((idx, row, all_issues))
            continue  # 未翻譯時不繼續做其他檢查

        # 術語檢查（已有譯文才做）
        eng_term = str(eng)
        trans_str = str(trans)
        glossary_issues = check_glossary_usage(eng_term, trans_str, relevant_glossary)
        for eng_term, chn, desc in glossary_issues:
            all_issues.append(("glossary", f"{eng_term} → {chn}", desc))

        # 佔位符檢查
        ph_issues = check_placeholder(eng_term, trans_str)
        if ph_issues:
            for pi in ph_issues:
                all_issues.append(("placeholder", "", pi))

        # 空格檢查
        space_issues = check_space_issues(trans_str)
        for _, _, desc in space_issues:
            all_issues.append(("space", "", desc))

        if all_issues:
            pool.append((idx, row, all_issues))

    return pool

def _filter_llm_pool(pool):
    """Keep only entries needing LLM retranslation; pure space issues are handled by scripts."""
    return [item for item in pool if any(itype != "space" for itype, _, _ in item[2])]


async def _retry_round(client, model_name, pool, glossary_text, rnd):
    """非同步執行一輪重譯，返回修正後的條目數"""
    # 每批 BATCH_SIZE_LIMIT（100）條，無群組機制
    all_results = []
    for bs in range(0, len(pool), BATCH_SIZE_LIMIT):
        batch = pool[bs:bs + BATCH_SIZE_LIMIT]
        items = [{"index": idx, "english": str(row["english"])} for idx, row, _ in batch]
        prompt = (RETRY_PROMPTS[rnd] + "\n" + json.dumps(items, ensure_ascii=False)
                  + f"\n## 强制术语表（请严格使用以下翻译）\n{glossary_text}"
                  + "\n请回传 JSON 阵列，每条包含 index 和 translation。")
        content = ""
        for attempt in range(3):
            try:
                resp = await client.chat.completions.create(
                    model=model_name,
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0, timeout=API_TIMEOUT,
                )
                # 保持正則操作，不要改為字串，否則會大幅降低翻譯成功率
                content = resp.choices[0].message.content.strip()
                content = re.sub(r"^```(?:json)?\n?", "", content, flags=re.IGNORECASE)
                content = re.sub(r"\n```$", "", content)
                parsed = json.loads(content)
                if len(parsed) > int(len(batch) * OVER_RETURN_TOLERANCE):
                    raise ValueError(f"回傳 {len(parsed)} 條，超過批次 {len(batch)} 的 {int(OVER_RETURN_TOLERANCE * 100)}%，視為失敗重試")
                all_results.extend(parsed)
                # 保持正則操作，不要改為字串，否則會大幅降低翻譯成功率
                break
            except json.JSONDecodeError as e:
                # ↓ 嘗試用 json_repair 修復
                try:
                    from json_repair import repair_json
                    repaired = repair_json(content)
                    if repaired:
                        fixed = json.loads(repaired)
                        if isinstance(fixed, list):
                            all_results.extend(fixed)
                            print(f"    ⚠️ JSON 修復成功")
                            break
                except Exception:
                    pass
                # ↑ 修復結束
                print(f"    JSON 解析失敗 (嘗試 {attempt + 1}/3): {e}")
                print(f"    原始回傳內容前 200 字: {content[:200]}")
                if attempt < 2:
                    await asyncio.sleep(2)
            except Exception as e:
                error_str = str(e).lower()
                if "429" in error_str or "rate" in error_str:
                    raise
                print(f"    API 錯誤 (嘗試 {attempt + 1}/3): {e}")
                if attempt < 2:
                    await asyncio.sleep(5)
        else:
            all_results.extend([{"index": idx, "translation": None, "_error": "API failed"}
                               for idx, _, _ in batch])
    batch_indices = {item[0] for item in pool}
    mapped = {}
    for r in all_results:
        if isinstance(r, dict) and r.get("index") in batch_indices:
            mapped[r["index"]] = r
    return list(mapped.values())

async def _enforce_async(df, relevant_glossary, glossary_text, output_dir=None, sheet_name=None, shared_pool=None):
    """非同步執行重譯循環，使用多 API 共享隊列模式（支援續傳）"""
    from llm_translator import save_backup_part, _sanitize_sheet_name
    import uuid
    if shared_pool is None:
        print("  無可用 API，跳過重譯")
        return df

    # ── 檢查 enforce checkpoint ──
    start_round = 0
    enforce_tag = uuid.uuid4().hex[:8]
    cp_file = None
    if output_dir and sheet_name:
        cp_file = Path(output_dir) / "_checkpoint" / _sanitize_sheet_name(sheet_name) / "enforce_checkpoint.json"
        if cp_file.exists():
            try:
                cp_data = json.loads(cp_file.read_text(encoding="utf-8"))
                start_round = cp_data.get("completed_rounds", 0)
                # 從 checkpoint 載入之前生成的 enforce_tag，若無則用新的
                enforce_tag = cp_data.get("enforce_tag", enforce_tag)
                print(f"  偵測到重譯進度，從第 {start_round + 1} 輪開始")
            except Exception:
                pass

    for rnd in range(start_round, 3):
        pool = scan_issues(df, relevant_glossary)


        # ★ 腳本預處理（每輪重譯前都執行）
        if pool:
            # 統計預處理前的問題類型
            glossary_cnt = sum(1 for _, _, issues in pool for t, _, _ in issues if t == "glossary")
            ph_cnt = sum(1 for _, _, issues in pool for t, _, _ in issues if t == "placeholder")
            untranslated_cnt = sum(1 for _, _, issues in pool for t, _, _ in issues if t == "untranslated")
            space_cnt = sum(1 for _, _, issues in pool for t, _, _ in issues if t == "space")
            before_parts = []
            if glossary_cnt > 0:
                before_parts.append(f"術語 {glossary_cnt} 條")
            if ph_cnt > 0:
                before_parts.append(f"佔位符 {ph_cnt} 條")
            if untranslated_cnt > 0:
                before_parts.append(f"未翻譯 {untranslated_cnt} 條")
            if space_cnt > 0:
                before_parts.append(f"空格 {space_cnt} 條")
            if before_parts:
                print(f"  ⚡ 腳本預處理前問題：{'、'.join(before_parts)}")

            preprocess_issues(df, pool)
            pool = scan_issues(df, relevant_glossary)

            # 統計修正了哪些類型
            space_after = sum(1 for _, _, issues in pool for t, _, _ in issues if t == "space")
            ph_after = sum(1 for _, _, issues in pool for t, _, _ in issues if t == "placeholder")
            solved_parts = []
            solved_space = space_cnt - space_after
            solved_ph = ph_cnt - ph_after
            if solved_space > 0:
                solved_parts.append(f"空格 {solved_space} 條")
            if solved_ph > 0:
                solved_parts.append(f"佔位符 {solved_ph} 條")
            if solved_parts:
                print(f"  ⚡ 腳本預處理修正：{'、'.join(solved_parts)}")
        if not pool:
            print(f"  第 {rnd + 1} 轮检查：全部正确")
            break
        llm_pool = _filter_llm_pool(pool)
        if not llm_pool:
            print(f"  第 {rnd + 1} 轮检查：剩餘問題皆為空格問題，由腳本處理，跳過重譯")
            break
        if len(llm_pool) < 3:
            print(f"  第 {rnd + 1} 轮检查：{len(llm_pool)} 条有问题（少于3条，跳过重译）")
            break
        pool = llm_pool

        # ── 掃描問題條目 ──
        gloss_count = sum(1 for _, _, issues in pool for t, _, _ in issues if t == "glossary")
        ph_count = sum(1 for _, _, issues in pool for t, _, _ in issues if t == "placeholder")
        untrans_count = sum(1 for _, _, issues in pool for t, _, _ in issues if t == "untranslated")
        space_count = sum(1 for _, _, issues in pool for t, _, _ in issues if t == "space")
        print(f"  ── 掃描問題條目 ──")
        print(f"    術語檢查：{gloss_count} 條")
        if ph_count > 0:
            print(f"    佔位符檢查：{ph_count} 條")
        if untrans_count > 0:
            print(f"    未翻譯檢查：{untrans_count} 條")
        if space_count > 0:
            print(f"    空格檢查：{space_count} 條")
        print(f"  第 {rnd + 1} 轮检查：{len(pool)} 条有问题，进行重譯...")

        pre_count = len(pool)

        # 分割問題條目為 BATCH_SIZE_LIMIT 大小的批次，丟進共享池
        batches = [pool[bs:bs + BATCH_SIZE_LIMIT] for bs in range(0, len(pool), BATCH_SIZE_LIMIT)]
        total_batches = len(batches)

        async def _retry_process(ws, job):
            cfg = ws["cfg"]
            batch_slice = job.batch
            try:
                results = await _retry_round(ws["client"], cfg.model, batch_slice, glossary_text, rnd)
                batch_indices = {item[0] for item in batch_slice}
                fixed = 0
                skipped = 0
                for res in results:
                    if not isinstance(res, dict):
                        skipped += 1
                        continue
                    idx = res.get("index")
                    trans = res.get("translation")
                    if idx is not None and trans is not None:
                        if idx in batch_indices:
                            df.at[idx, "translation"] = trans
                            fixed += 1
                        else:
                            skipped += 1
                if skipped:
                    print(f"  [{cfg.api_id}] 警告：{skipped} 條回傳無效（非物件或 index 不在工作表中），已忽略")
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                print(f"  [{now}] [重譯] 批次 {job.batch_num}/{total_batches} 完成 ({fixed} 条已修正) [{cfg.api_id} {cfg.model}]")
            except Exception as e:
                error_str = str(e).lower()
                if "429" in error_str or "rate" in error_str:
                    cfg.mark_429()
                    if cfg.is_permanently_disabled:
                        print(f"  [{cfg.api_id} {cfg.model}] 第 {cfg.strike} 次 429，永久停用")
                    else:
                        print(f"  [{cfg.api_id} {cfg.model}] 第 {cfg.strike} 次 429，冷卻 60 秒")
                    shared_pool.retry(job)
                else:
                    print(f"  [{cfg.api_id} {cfg.model}] 錯誤：{e}")

        if batches:
            await asyncio.gather(*[shared_pool.submit(bn, b, _retry_process) for bn, b in enumerate(batches, 1)])

        # 計算修正率，決定是否繼續
        post_pool = scan_issues(df, relevant_glossary)

        # ★ 儲存本輪修正到 part 檔案（原子寫入） ★
        if output_dir:
            backup_data = {}
            for idx, row, _ in pool:
                trans = df.at[idx, "translation"]
                if trans is not None:
                    backup_data[str(idx)] = str(trans)
            if backup_data:
                save_backup_part(output_dir, enforce_tag, rnd + 1, backup_data, sheet_name)

        # ★ 儲存 enforce checkpoint ★
        if cp_file:
            cp_file.parent.mkdir(parents=True, exist_ok=True)
            atomic_write_text(cp_file, json.dumps({
                "completed_rounds": rnd + 1,
                "enforce_tag": enforce_tag,
                "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
            }, ensure_ascii=False, indent=2))

        post_all_count = len(post_pool)
        post_count = len(_filter_llm_pool(post_pool))  # 同口徑：與送重譯前一致（不含純空格）
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if post_all_count == 0:
            print(f"  [{now}] 第 {rnd + 1} 轮重译后：全部正确")
            break
        if post_count == 0:
            print(f"  [{now}] 第 {rnd + 1} 轮重译后：剩餘問題皆為空格問題，由腳本處理")
            break

        corrected = pre_count - post_count
        rate = corrected / pre_count if pre_count > 0 else 0
        print(f"  [{now}] 第 {rnd + 1} 轮重译后：剩余 {post_count} 条，修正率 {rate * 100:.1f}%")
        if rate < 0.05:
            print(f"  [{now}] 修正率低於 5%，跳過後續輪次")
            break

    # ── 所有輪次完成，刪除 enforce checkpoint ──
    if cp_file and cp_file.exists():
        cp_file.unlink()
        print("  重譯完成，已清除進度記錄")

    return df

async def enforce_async(df: pd.DataFrame, glossary: dict, output_dir: str | Path | None = None,
                        report_name: str = "review_report.xlsx", write_report: bool = True,
                        sheet_name: str | None = None, shared_pool=None) -> tuple:
    df = df.copy()
    review_rows = []

    # 只保留對當前批次有相關性的術語
    all_text = " ".join(str(row.get("english", "")).lower() for _, row in df.iterrows())
    relevant_glossary = {
        e: c for e, c in glossary.items()
        if re.search(r"(?<![a-z'])" + re.escape(e.lower()) + r"(?![a-z'])", all_text)
    }
    glossary_text = "\n".join([f"  {e} → {c}" for e, c in relevant_glossary.items()]) if relevant_glossary else "  无"

    # 非同步重譯（使用共享池；未提供時建立臨時池）
    if shared_pool is not None:
        df = await _enforce_async(df, relevant_glossary, glossary_text, output_dir, sheet_name, shared_pool)
    else:
        tmp_pool = SharedBatchPool()
        try:
            await tmp_pool.start()
        except RuntimeError:
            tmp_pool = None
        if tmp_pool is not None:
            try:
                df = await _enforce_async(df, relevant_glossary, glossary_text, output_dir, sheet_name, tmp_pool)
            finally:
                await tmp_pool.close()

    # 最終審查掃描
    final_pool = scan_issues(df, relevant_glossary)
    for idx, row, issues in final_pool:
        trans = row.get("translation")
        trans_str = str(trans) if not (trans is None or (isinstance(trans, float) and pd.isna(trans))) else ""
        for issue_type, suggested_fix, desc in issues:
            review_rows.append({
                "english": str(row.get("english", "")),
                "current_translation": trans_str,
                "category": str(row.get("category", "")),
                "issue": f"[{issue_type}] {desc}",
            })

    review_df = pd.DataFrame(review_rows)
    if write_report and not review_df.empty:
        if not output_dir:
            print("  警告：未提供 output_dir，跳過審查報告寫入")
            return df, review_df
        fp = Path(output_dir) / report_name

        from openpyxl.styles import Color, PatternFill

        fill_glossary = PatternFill(start_color=Color(rgb="D9D9D9"), end_color=Color(rgb="D9D9D9"), fill_type="solid")
        fill_placeholder = PatternFill(start_color=Color(rgb="BDD7EE"), end_color=Color(rgb="BDD7EE"), fill_type="solid")
        fill_untranslated = PatternFill(start_color=Color(rgb="F4CCCC"), end_color=Color(rgb="F4CCCC"), fill_type="solid")
        fill_space = PatternFill(start_color=Color(rgb="E2EFDA"), end_color=Color(rgb="E2EFDA"), fill_type="solid")

        with pd.ExcelWriter(fp, engine='openpyxl') as writer:
            review_df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']

            # 找到 issue 欄位的欄號
            issue_col = None
            for col_idx, cell in enumerate(worksheet[1], start=1):
                if cell.value == "issue":
                    issue_col = col_idx
                    break

            if issue_col:
                for row in worksheet.iter_rows(min_row=2, max_col=worksheet.max_column, max_row=worksheet.max_row):
                    issue_cell = row[issue_col - 1]
                    issue_val = str(issue_cell.value) if issue_cell.value else ""
                    if issue_val.startswith("[glossary]"):
                        fill = fill_glossary
                    elif issue_val.startswith("[placeholder]"):
                        fill = fill_placeholder
                    elif issue_val.startswith("[untranslated]"):
                        fill = fill_untranslated
                    elif issue_val.startswith("[space]"):
                        fill = fill_space
                    else:
                        continue
                    for cell in row:
                        cell.fill = fill

        print(f"  审查报告已写入: {fp} ({len(review_df)} 条需审查)")
    return df, review_df


